"""Refresh audit/model_outputs.json from the live Excel model.

Pulls every value the deck and visuals need, in one pass, deterministically.
Run this whenever model/WHOOP Valuation Model Update.xlsx changes before
running build_visuals.py and build_deck.py.
"""
from __future__ import annotations
from openpyxl import load_workbook
from pathlib import Path
import json

MODEL = Path("/home/user/whoop_val/model/WHOOP Valuation Model Update.xlsx")
OUT   = Path("/home/user/whoop_val/audit/model_outputs.json")

wb = load_workbook(MODEL, data_only=True)


def cell(sheet: str, addr: str):
    return wb[sheet][addr].value


def num(v):
    if v in (None, "", "n/a", "—"): return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(v)
    except Exception: return v


def row_dict(sheet: str, start_row: int, key_col: int, value_col: int, n_rows: int) -> dict:
    ws = wb[sheet]
    out = {}
    for i in range(n_rows):
        k = ws.cell(row=start_row + i, column=key_col).value
        v = ws.cell(row=start_row + i, column=value_col).value
        if k is not None:
            out[str(k)] = num(v)
    return out


# =============================================================
# DCF core: per-scenario equity, scen-weighted, prob-weighted
# =============================================================
dcf = wb["DCF"]
ev_base   = num(cell("DCF", "C53"))   # legacy fallback — confirm below
# We rely on the Scenarios sheet which is authoritative for cross-scenario links
sc = wb["Scenarios"]
# Row 5 = Intrinsic DCF, columns B (Bear) C (Base) D (Bull) E (ProbWtd)
bear_equity = num(sc["B5"].value)
base_equity = num(sc["C5"].value)
bull_equity = num(sc["D5"].value)
pw_equity   = num(sc["E5"].value)

# Rows 16-18 — three weight schemes
schemes = {}
for r, name in [(16, "pessimistic"), (17, "neutral"), (18, "optimistic")]:
    schemes[name] = {
        "p_bear":    num(sc.cell(row=r, column=2).value),
        "p_base":    num(sc.cell(row=r, column=3).value),
        "p_bull":    num(sc.cell(row=r, column=4).value),
        "scen_wtd":  num(sc.cell(row=r, column=5).value),
        "prob_wtd":  num(sc.cell(row=r, column=6).value),
        "jensen":    num(sc.cell(row=r, column=7).value),
        "vs_serg":   num(sc.cell(row=r, column=8).value),
    }

# Legacy single-value pointers (kept for backward compatibility with build_deck.py)
scen_pess = schemes["pessimistic"]["scen_wtd"]
scen_neu  = schemes["neutral"]["scen_wtd"]
scen_opt  = schemes["optimistic"]["scen_wtd"]
input_pess = schemes["pessimistic"]["prob_wtd"]
input_neu  = schemes["neutral"]["prob_wtd"]
input_opt  = schemes["optimistic"]["prob_wtd"]


# =============================================================
# Football Field — methods rows 6-14 + chart helper
# =============================================================
ff = wb["Football Field"]
ff_rows = {}
for r in [6, 8, 9, 10, 12, 13, 14]:
    method = ff.cell(row=r, column=1).value
    if method:
        ff_rows[f"row_{r}"] = {
            "method": method,
            "low":    num(ff.cell(row=r, column=3).value),
            "base":   num(ff.cell(row=r, column=4).value),
            "high":   num(ff.cell(row=r, column=5).value),
        }
ff_weighted_avg = num(ff["D18"].value)            # weighted average EV ($B)
ff_prem_serg    = num(ff["D20"].value)            # premium to Series G

# IPO range — anchored on private→IPO step-up reference class
ipo_bear = num(ff["D61"].value)
ipo_base = num(ff["D62"].value)
ipo_bull = num(ff["D63"].value)
# Reference deals
ipo_refdeals = []
for r in range(46, 52):
    ipo_refdeals.append({
        "company":    ff.cell(row=r, column=1).value,
        "step_up":    num(ff.cell(row=r, column=8).value),
    })


# =============================================================
# Section 14 — Assumption Inventory (Assumptions R215-R235)
# =============================================================
a = wb["Assumptions"]
inventory = []
for r in range(216, 236):
    section = a.cell(row=r, column=1).value
    driver  = a.cell(row=r, column=2).value
    if not driver: continue
    inventory.append({
        "section":  section,
        "driver":   driver,
        "bear":     num(a.cell(row=r, column=3).value),
        "base":     num(a.cell(row=r, column=4).value),
        "bull":     num(a.cell(row=r, column=5).value),
        "weighted": num(a.cell(row=r, column=6).value),
        "active":   num(a.cell(row=r, column=10).value),
        "spread":   num(a.cell(row=r, column=11).value),
    })


# =============================================================
# Grid 6 — P_Bear × P_Bull heatmap (Sensitivity R65-R70)
# =============================================================
sens = wb["Sensitivity"]
# Header row 65: P_Bear↓ / P_Bull→ then 0.20, 0.25, 0.30, 0.35, 0.40
p_bull_axis = [num(sens.cell(row=65, column=c).value) for c in range(2, 7)]
grid6 = {"p_bull_axis": p_bull_axis, "rows": []}
for r in range(66, 71):
    p_bear = num(sens.cell(row=r, column=1).value)
    vals = [num(sens.cell(row=r, column=c).value) for c in range(2, 7)]
    grid6["rows"].append({"p_bear": p_bear, "equity_b": vals})


# Grid 7 — Members lens
grid7 = []
for r in range(78, 84):
    grid7.append({
        "lens":         sens.cell(row=r, column=1).value,
        "members_2033": num(sens.cell(row=r, column=2).value),
        "revenue_2033": num(sens.cell(row=r, column=3).value),
        "ev_const":     num(sens.cell(row=r, column=4).value),
        "ev_coupled":   num(sens.cell(row=r, column=5).value),
        "sh_coupled":   num(sens.cell(row=r, column=7).value),
    })


# Grid 1 — WACC × Exit Multiple (Sensitivity R5-R12)
grid1_x = [num(sens.cell(row=5, column=c).value) for c in range(2, 9)]
grid1 = {"exit_mult_axis": grid1_x, "rows": []}
for r in range(6, 13):
    wacc = num(sens.cell(row=r, column=1).value)
    vals = [num(sens.cell(row=r, column=c).value) for c in range(2, 9)]
    grid1["rows"].append({"wacc": wacc, "equity_b": vals})


# =============================================================
# Comps table (Comps R19-R29)
# =============================================================
c = wb["Comps"]
comps_full = []
for r in range(19, 30):
    name = c.cell(row=r, column=1).value
    if not name: continue
    comps_full.append({
        "name":      name,
        "ticker":    c.cell(row=r, column=2).value,
        "tier":      c.cell(row=r, column=3).value,
        "thesis":    c.cell(row=r, column=4).value,
        "ev_rev":    num(c.cell(row=r, column=5).value),
        "ev_gp":     num(c.cell(row=r, column=6).value),
        "growth":    num(c.cell(row=r, column=7).value),
        "gm":        num(c.cell(row=r, column=8).value),
    })

# Thesis ranges
thesis = {
    "t1_failed_pivot":  {"low": num(c["B39"].value), "high": num(c["C39"].value), "median": num(c["D39"].value)},
    "t2_sub_executed":  {"low": num(c["B50"].value), "high": num(c["C50"].value), "median": num(c["D50"].value)},
    "t3_healthcare":    {"low": num(c["B61"].value), "high": num(c["C61"].value), "median": num(c["D61"].value)},
}


# =============================================================
# Legacy fields — pulled from existing JSON to preserve compatibility
# =============================================================
LEGACY_JSON = Path("/home/user/whoop_val/audit/model_outputs.json")
legacy = json.loads(LEGACY_JSON.read_text()) if LEGACY_JSON.exists() else {}


# Compose the output
output = {
    # ----- legacy keys (preserved with refreshed numbers) -----
    "series_g":        10100.0,
    "bear_equity":     bear_equity * 1000 if bear_equity and bear_equity < 100 else bear_equity,
    "base_equity":     base_equity * 1000 if base_equity and base_equity < 100 else base_equity,
    "bull_equity":     bull_equity * 1000 if bull_equity and bull_equity < 100 else bull_equity,
    "pw_equity":       pw_equity   * 1000 if pw_equity   and pw_equity   < 100 else pw_equity,
    "scen_pess":       scen_pess   * 1000 if scen_pess   and scen_pess   < 100 else scen_pess,
    "scen_neu":        scen_neu    * 1000 if scen_neu    and scen_neu    < 100 else scen_neu,
    "scen_opt":        scen_opt    * 1000 if scen_opt    and scen_opt    < 100 else scen_opt,
    "input_pess":      input_pess  * 1000 if input_pess  and input_pess  < 100 else input_pess,
    "input_neu":       input_neu   * 1000 if input_neu   and input_neu   < 100 else input_neu,
    "input_opt":       input_opt   * 1000 if input_opt   and input_opt   < 100 else input_opt,
    "ff_d18":          ff_weighted_avg,
    "ff_d20":          ff_prem_serg,
    "ff_d6":           ff_rows["row_6"]["base"],
    "ff_d9":           ff_rows["row_9"]["base"],
    "ff_d10":          ff_rows["row_10"]["base"],
    "ff_d13":          ff_rows["row_13"]["base"],
    "ff_d14":          ff_rows["row_14"]["base"],
    # ----- new keys -----
    "ff_rows":         ff_rows,
    "ipo_range":       {"bear": ipo_bear, "base": ipo_base, "bull": ipo_bull, "deals": ipo_refdeals},
    "weight_schemes":  schemes,
    "inventory":       inventory,
    "grid1":           grid1,
    "grid6":           grid6,
    "grid7":           grid7,
    "comps_full":      comps_full,
    "thesis_ranges":   thesis,
}

# Carry over keys from legacy JSON that we didn't recompute
for k, v in legacy.items():
    if k not in output:
        output[k] = v

OUT.write_text(json.dumps(output, indent=2, default=str))
print(f"Wrote {OUT}: {len(json.dumps(output))} bytes, {len(output)} top-level keys")

# Sanity checks
print("\n--- Sanity ---")
print(f"  series_g:              ${output['series_g']:,.0f}M")
print(f"  base_equity:           ${output['base_equity']:,.1f}M")
print(f"  scen_neu (Neutral):    ${output['scen_neu']:,.1f}M  (was 11556.93)")
print(f"  ff_d14 (IPO Base):     ${output['ff_d14']:.2f}B   (target ~$14.69B)")
print(f"  ff_d18 (FF Wtd Avg):   ${output['ff_d18']:.2f}B   (target ~$9.09B)")
print(f"  inventory drivers:     {len(inventory)} rows")
print(f"  grid6 rows × cols:     {len(grid6['rows'])} × {len(grid6['p_bull_axis'])}")
print(f"  comps_full:            {len(comps_full)} comps")
