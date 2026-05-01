#!/usr/bin/env python3
"""
WHOOP Master Valuation Model — Workbook Builder
Generates whoop-master-model.xlsx with IB-standard formatting.
All deterministic math done in Excel formulas, not Python.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Style Constants ──────────────────────────────────────────────────────────

# IB font conventions
FONT_INPUT = Font(name='Calibri', size=10, color='0000FF', bold=False)       # Blue = hard-coded input
FONT_FORMULA = Font(name='Calibri', size=10, color='000000', bold=False)     # Black = formula
FONT_LINK = Font(name='Calibri', size=10, color='008000', bold=False)        # Green = cross-sheet link
FONT_HEADER = Font(name='Calibri', size=10, color='000000', bold=True)
FONT_SECTION = Font(name='Calibri', size=11, color='000000', bold=True)
FONT_TITLE = Font(name='Calibri', size=14, color='000000', bold=True)
FONT_SUBTITLE = Font(name='Calibri', size=11, color='666666', bold=False)
FONT_UNITS = Font(name='Calibri', size=9, color='999999', italic=True)
FONT_FLAG = Font(name='Calibri', size=10, color='CC0000', bold=False)        # Red = flagged gap
FONT_NOTE = Font(name='Calibri', size=9, color='666666', italic=True)

# Fills
FILL_HEADER = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Light blue header
FILL_SECTION = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Light green section
FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Yellow = needs verification
FILL_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')     # Gray section divider
FILL_INPUT = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')    # Light green input cell
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Border
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)
BOTTOM_BORDER = Border(
    bottom=Side(style='medium', color='000000')
)

# Alignment
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Number formats
FMT_MILLIONS = '#,##0'
FMT_MILLIONS_1 = '#,##0.0'
FMT_PCT = '0.0%'
FMT_PCT_0 = '0%'
FMT_MULTIPLE = '0.0x'
FMT_DOLLARS = '$#,##0'
FMT_DOLLARS_M = '$#,##0'
FMT_SHARE_PRICE = '$#,##0.00'


def set_col_widths(ws, widths):
    """Set column widths from a dict of {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_row(ws, row, data, fonts=None, fills=None, formats=None, aligns=None):
    """Write a row of data with optional styling per cell."""
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=i+1, value=val)
        if fonts and i < len(fonts) and fonts[i]:
            cell.font = fonts[i]
        if fills and i < len(fills) and fills[i]:
            cell.fill = fills[i]
        if formats and i < len(formats) and formats[i]:
            cell.number_format = formats[i]
        if aligns and i < len(aligns) and aligns[i]:
            cell.alignment = aligns[i]


def section_header(ws, row, text, num_cols=15):
    """Write a gray section header spanning multiple columns."""
    for c in range(1, num_cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_GRAY
        cell.font = FONT_SECTION
        if c == 1:
            cell.value = text


def col_headers(ws, row, headers, num_fmt_row=None):
    """Write column headers with header fill."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BOTTOM_BORDER


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: COVER
# ═══════════════════════════════════════════════════════════════════════════════

def build_cover(wb):
    ws = wb.active
    ws.title = 'Cover'
    ws.sheet_properties.tabColor = '1F4E79'
    set_col_widths(ws, {'A': 5, 'B': 50, 'C': 30})

    ws.cell(row=3, column=2, value='WHOOP Inc.').font = Font(name='Calibri', size=24, bold=True)
    ws.cell(row=4, column=2, value='Private Company Valuation').font = Font(name='Calibri', size=16, color='666666')
    ws.cell(row=5, column=2, value='Investment Committee Memo — Supporting Model').font = FONT_SUBTITLE

    ws.cell(row=8, column=2, value='Company:').font = FONT_HEADER
    ws.cell(row=8, column=3, value='WHOOP Inc.').font = FONT_INPUT
    ws.cell(row=9, column=2, value='Ticker:').font = FONT_HEADER
    ws.cell(row=9, column=3, value='Private (Pre-IPO)').font = FONT_INPUT
    ws.cell(row=10, column=2, value='Last Round:').font = FONT_HEADER
    ws.cell(row=10, column=3, value='Series G — $575M at $10.1B post-money (Mar 31, 2026)').font = FONT_INPUT
    ws.cell(row=11, column=2, value='Model Date:').font = FONT_HEADER
    ws.cell(row=11, column=3, value='April 16, 2026').font = FONT_INPUT
    ws.cell(row=12, column=2, value='Analyst:').font = FONT_HEADER
    ws.cell(row=12, column=3, value='Landon Brice').font = FONT_INPUT
    ws.cell(row=13, column=2, value='Fiscal Year:').font = FONT_HEADER
    ws.cell(row=13, column=3, value='Calendar Year (Jan-Dec)').font = FONT_INPUT
    ws.cell(row=14, column=2, value='Currency:').font = FONT_HEADER
    ws.cell(row=14, column=3, value='USD ($M unless noted)').font = FONT_INPUT

    ws.cell(row=17, column=2, value='Valuation Methods:').font = FONT_SECTION
    methods = [
        '1. Intrinsic DCF (scenario-based + real options)',
        '2. Trading Comps (three-bucket framework)',
        '3. Precedent Transactions',
        '4. Last-Round Implied ($10.1B)',
        '5. Implied IPO Trading Range (back-solved)',
    ]
    for i, m in enumerate(methods):
        ws.cell(row=18+i, column=2, value=m).font = FONT_FORMULA

    ws.cell(row=25, column=2, value='Color Convention:').font = FONT_SECTION
    ws.cell(row=26, column=2, value='Hard-coded input').font = FONT_INPUT
    ws.cell(row=27, column=2, value='Formula / calculation').font = FONT_FORMULA
    ws.cell(row=28, column=2, value='Cross-sheet link').font = FONT_LINK
    c = ws.cell(row=29, column=2, value='Needs verification / FLAGGED_GAP')
    c.font = FONT_FLAG
    c.fill = FILL_YELLOW

    ws.cell(row=32, column=2, value='Source Files:').font = FONT_SECTION
    ws.cell(row=33, column=2, value='See Sources tab for full research file index').font = FONT_NOTE


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def build_assumptions(wb):
    ws = wb.create_sheet('Assumptions')
    ws.sheet_properties.tabColor = '4472C4'
    set_col_widths(ws, {
        'A': 8, 'B': 35, 'C': 18, 'D': 18, 'E': 18, 'F': 12, 'G': 12, 'H': 40
    })

    ws.cell(row=1, column=1, value='WHOOP — Master Assumptions').font = FONT_TITLE

    # Scenario toggle
    ws.cell(row=3, column=1, value='Scenario Toggle:').font = FONT_SECTION
    toggle_cell = ws.cell(row=3, column=2, value='Base')
    toggle_cell.font = Font(name='Calibri', size=12, color='0000FF', bold=True)
    toggle_cell.fill = FILL_INPUT
    ws.cell(row=3, column=3, value='← Enter: Bear / Base / Bull').font = FONT_NOTE
    # Name the toggle cell for formulas
    from openpyxl.workbook.defined_name import DefinedName
    dn = DefinedName('Scenario', attr_text="Assumptions!$B$3")
    wb.defined_names.add(dn)

    r = 5
    col_headers(ws, r, ['ID', 'Assumption', 'Bear', 'Base', 'Bull', 'Tier', 'Priority', 'Source / Notes'])
    ws.cell(row=r+1, column=3, value='').font = FONT_UNITS  # units row placeholder

    r = 7
    # ── Section 1: Revenue & Growth ──
    section_header(ws, r, 'Section 1: Revenue & Growth')
    r += 1

    assumptions_rev = [
        ['R1', '2025 Recognized Revenue ($M)', 550, 650, 800, 'Tier 2', 'Critical', 'Stress-tested range; $1.1B is bookings run-rate not recognized rev'],
        ['R2', '2025 Bookings Run-Rate ($M)', 1100, 1100, 1100, 'Tier 1', 'Locked', 'Series G press release; CEO statement'],
        ['R3', 'Paying Members (M)', 2.3, 2.5, 2.7, 'Tier 1', 'Critical', 'Series G press: 2.5M+'],
        ['R4', 'Blended ARPU ($/yr)', 280, 302, 330, 'Tier 3', 'Critical', 'Bottom-up model; revenue-bookings-reconciliation.md'],
        ['R5', 'Subscription Revenue %', 0.80, 0.85, 0.88, 'Tier 2', 'High', 'Sacra research; 80-88% range'],
        ['R6', '2025 YoY Subscription Growth', 1.03, 1.03, 1.03, 'Tier 1', 'Locked', 'Series G press release'],
        ['R7', 'Phase 1 Revenue Growth (2026-27)', 0.50, 0.70, 0.90, 'Tier 3', 'Critical', 'Peer trajectories; spec D phase table'],
        ['R8', 'Phase 2 Revenue Growth (2028-30)', 0.15, 0.25, 0.40, 'Tier 3', 'Critical', 'Deceleration curves from Dexcom/Netflix'],
        ['R9', 'Phase 3 Revenue Growth (2031-33)', 0.05, 0.12, 0.20, 'Tier 3', 'High', 'TAM ceiling, competitive response'],
        ['R10', 'Annual Churn Rate', 0.22, 0.17, 0.12, 'Tier 3', 'Critical', '>80% retention = <20% churn; FLAGGED_GAP: no cohort data'],
        ['R11', 'International Revenue %', 0.35, 0.43, 0.55, 'Tier 2', 'High', 'Stress-tested: 40-45% base; NYT/DealBook source ambiguous'],
        ['R12', 'B2B Unite Revenue %', 0.12, 0.18, 0.25, 'Tier 2', 'Medium', 'Sacra/BMCT: ~20% of recurring; growing rapidly'],
    ]

    for a in assumptions_rev:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float) and val < 1.5:
                    cell.number_format = FMT_PCT
                else:
                    cell.number_format = FMT_MILLIONS
            elif ci == 5:
                cell.font = FONT_NOTE
            elif ci == 6:
                cell.font = FONT_HEADER if val == 'Critical' else FONT_NOTE
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        # Flag unsourced assumptions
        if a[5] == 'Tier 3' and a[6] == 'Critical':
            for ci in range(8):
                ws.cell(row=r, column=ci+1).fill = FILL_YELLOW
        r += 1

    r += 1
    # ── Section 2: Cost Structure & Margins ──
    section_header(ws, r, 'Section 2: Cost Structure & Margins')
    r += 1

    assumptions_cost = [
        ['C1', 'Hardware BOM per Device ($)', 100, 80, 60, 'Tier 3', 'Critical', 'Est. $50-80 at scale; 2016 was $250-300'],
        ['C2', 'Subscription Gross Margin', 0.72, 0.78, 0.83, 'Tier 3', 'Critical', 'Peer: SaaS ~80%, Peloton sub ~68%'],
        ['C3', 'Hardware Gross Margin', -0.10, -0.02, 0.05, 'Tier 3', 'Critical', 'Subsidized; Peloton HW GM: -18% to +14%'],
        ['C4', 'Blended GM Phase 1 (2026-27)', 0.56, 0.60, 0.64, 'Tier 3', 'Critical', 'Derived from C2+C3+growth rate'],
        ['C5', 'Blended GM Phase 2 (2028-30)', 0.63, 0.67, 0.72, 'Tier 3', 'High', 'Margin expansion as growth decelerates'],
        ['C6', 'Blended GM Terminal', 0.72, 0.76, 0.80, 'Tier 3', 'High', 'Peer terminal: Dexcom 60%, Spotify 32%'],
        ['C7', 'R&D % of Revenue (current)', 0.25, 0.22, 0.20, 'Tier 3', 'High', 'Peers: Dexcom 13%, Garmin 16%, growth SaaS 20-30%'],
        ['C8', 'R&D % of Revenue (terminal)', 0.15, 0.13, 0.12, 'Tier 3', 'Medium', 'Peer steady-state: Garmin 16%, Dexcom 13%'],
        ['C9', 'S&M % of Revenue (current)', 0.35, 0.30, 0.25, 'Tier 3', 'Critical', 'Peloton peak 28.4%; current 16.9%'],
        ['C10', 'S&M % of Revenue (terminal)', 0.20, 0.17, 0.15, 'Tier 3', 'High', 'Peer steady-state: Garmin SG&A 14-19%'],
        ['C11', 'G&A % of Revenue (current)', 0.18, 0.15, 0.12, 'Tier 3', 'Medium', 'IPO step-up expected 2027-28'],
        ['C12', 'G&A % of Revenue (terminal)', 0.12, 0.10, 0.08, 'Tier 3', 'Medium', 'Peer steady-state'],
        ['C13', 'SBC % of Revenue', 0.15, 0.12, 0.10, 'Tier 3', 'High', 'Peloton: 9.2% (FY2025); pre-IPO typically higher'],
        ['C14', 'CapEx % of Revenue', 0.05, 0.04, 0.03, 'Tier 3', 'Medium', 'Asset-light; contract manufacturing'],
        ['C15', 'Tax Rate (effective)', 0.25, 0.23, 0.21, 'Tier 2', 'Medium', 'US statutory; intl mix-shift lowers over time'],
    ]

    for a in assumptions_cost:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float) and -0.5 < val < 1.5:
                    cell.number_format = FMT_PCT
                else:
                    cell.number_format = FMT_DOLLARS
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        if a[5] == 'Tier 3' and a[6] == 'Critical':
            for ci in range(8):
                ws.cell(row=r, column=ci+1).fill = FILL_YELLOW
        r += 1

    r += 1
    # ── Section 3: WACC / Discount Rate ──
    section_header(ws, r, 'Section 3: WACC / Discount Rate (Sourced — Damodaran Apr 2026)')
    r += 1

    wacc_inputs = [
        ['D1', 'Risk-Free Rate (10yr UST)', 0.043, 0.043, 0.043, 'Tier 1', 'Point Est.', 'Fed H.15, Apr 14 2026: 4.26%; rounded 4.3%'],
        ['D2', 'Equity Risk Premium', 0.045, 0.045, 0.045, 'Tier 1', 'Point Est.', 'Damodaran 4.23% (Jan 2026); Kroll 5.0%; use 4.5%'],
        ['D3', 'Unlevered Beta (bucket-wtd)', 0.85, 0.83, 0.81, 'Tier 1', 'Sensitivity', 'Damodaran sector: Healthcare 0.83, Electronics 0.83'],
        ['D4', 'Size Premium', 0.0081, 0.0081, 0.0081, 'Tier 2', 'Point Est.', 'Kroll CRSP Decile 3 ($7.3-13.5B): 0.81%'],
        ['D5', 'Company-Specific Risk Premium', 0.03, 0.02, 0.01, 'Tier 3', 'Sensitivity', 'Private company illiquidity; 1-3% range'],
        ['D6', 'Implied Cost of Equity / WACC', None, None, None, 'Derived', 'Formula', '= D1 + D2×D3 + D4 + D5; ~9.5-12.0%'],
    ]

    for a in wacc_inputs:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if val is not None and isinstance(val, float):
                    cell.number_format = FMT_PCT
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1

    r += 1
    # ── Section 4: Terminal Value ──
    section_header(ws, r, 'Section 4: Terminal Value')
    r += 1

    tv_inputs = [
        ['D9', 'Terminal Method', 'Exit Multiple', 'Exit Multiple', 'Exit Multiple', 'Design', 'Critical', 'Primary: EV/Revenue exit multiple; cross-check: Gordon Growth'],
        ['D10', 'Terminal EV/Revenue Multiple', 3.5, 4.5, 6.0, 'Tier 3', 'Critical', 'Compressed from current comps; reflects terminal growth of 10-15%'],
        ['D11', 'Terminal EV/EBITDA Multiple', 12.0, 15.0, 18.0, 'Tier 3', 'High', 'Cross-check only; more stable at maturity'],
        ['D12', 'Perpetuity Growth Rate (Gordon)', 0.025, 0.03, 0.035, 'Tier 2', 'Critical', 'GDP growth + inflation; cannot exceed WACC'],
    ]

    for a in tv_inputs:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float) and val < 1.0:
                    cell.number_format = FMT_PCT
                elif isinstance(val, (int, float)):
                    cell.number_format = '0.0x'
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1

    r += 1
    # ── Section 5: Bucket Weighting ──
    section_header(ws, r, 'Section 5: Comp Bucket Weighting (THIS IS THE THESIS)')
    r += 1

    bucket_inputs = [
        ['V11a', 'Bucket 1 Weight (Hardware)', 0.40, 0.20, 0.10, 'Design', 'Critical', 'Garmin anchor; higher = more bearish'],
        ['V11b', 'Bucket 2 Weight (Subscription)', 0.40, 0.40, 0.30, 'Design', 'Critical', 'Peloton-Spotify range; WHOOP sits where?'],
        ['V11c', 'Bucket 3 Weight (Health Data)', 0.20, 0.40, 0.60, 'Design', 'Critical', 'Higher = healthcare thesis is believed'],
        ['V12', 'Implied Blended Multiple', None, None, None, 'Derived', 'Formula', '= Σ(weight × bucket median multiple); see Comps tab'],
    ]

    for a in bucket_inputs:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float):
                    cell.number_format = FMT_PCT
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1

    r += 1
    # ── Section 6: DLOM / IPO ──
    section_header(ws, r, 'Section 6: Discount & IPO Assumptions')
    r += 1

    dlom_inputs = [
        ['I1', 'IPO Timing (months from Series G)', 18, 15, 12, 'Tier 2', 'High', 'Q2 2027 - Q1 2028; "last private round"'],
        ['I2', 'IPO Discount to Fair Value', 0.25, 0.20, 0.15, 'Tier 2', 'Critical', 'Loughran/Ritter IPO studies'],
        ['I3', 'Illiquidity / Timing Discount', 0.15, 0.10, 0.05, 'Tier 2', 'High', 'Oura tender at 25% discount; Hiive 39-51%'],
        ['S9', 'DLOM (base case)', 0.30, 0.22, 0.15, 'Tier 2', 'High', 'Oura tender 25%; EquityZen avg 13-29%'],
        ['I4', 'Series G Investor Required IRR', 0.30, 0.25, 0.20, 'Tier 3', 'Medium', 'Late-stage growth equity return expectations'],
    ]

    for a in dlom_inputs:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float) and val < 1.0:
                    cell.number_format = FMT_PCT
                else:
                    cell.number_format = FMT_MILLIONS
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1

    r += 1
    # ── Section 7: Real Options ──
    section_header(ws, r, 'Section 7: Real Options (Additive — NEVER in DCF base case)')
    r += 1

    option_inputs = [
        ['D19', 'Healthcare Reimbursement', 0.20, 0.25, 0.30, 'Probability', '—', 'CMS Innovation Center selection upgrades this'],
        ['D20', 'International Expansion (incremental)', 0.30, 0.45, 0.60, 'Probability', '—', 'Partially in base (~40% intl already); incremental only'],
        ['D21', 'B2B Enterprise / Clinical', 0.20, 0.30, 0.40, 'Probability', '—', 'Unite already ~20%; incremental only'],
        ['D22', "Women's Health / Fertility", 0.25, 0.40, 0.50, 'Probability', '—', '150% YoY women membership growth; Apr 2026 panel launch'],
        ['D23', 'Noninvasive Glucose Monitoring', 0.03, 0.05, 0.07, 'Probability', '—', 'Deliberately conservative; hardest unsolved problem'],
    ]

    for a in option_inputs:
        for ci, val in enumerate(a):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci == 1:
                cell.font = FONT_FORMULA
            elif ci in (2, 3, 4):
                cell.font = FONT_INPUT
                if isinstance(val, float):
                    cell.number_format = FMT_PCT
            elif ci == 7:
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3: REVENUE BUILD
# ═══════════════════════════════════════════════════════════════════════════════

def build_revenue(wb):
    ws = wb.create_sheet('Revenue Build')
    ws.sheet_properties.tabColor = '548235'
    cols = {'A': 30, 'B': 5}
    for i, yr in enumerate(range(2024, 2034)):
        cols[get_column_letter(i+3)] = 14
    set_col_widths(ws, cols)

    ws.cell(row=1, column=1, value='WHOOP — Revenue Build-Up').font = FONT_TITLE
    ws.cell(row=2, column=1, value='$M unless noted; Calendar Year').font = FONT_UNITS

    years = list(range(2024, 2034))
    yr_labels = [''] + ['Units'] + [str(y) for y in years]

    r = 4
    col_headers(ws, r, yr_labels)

    # Phase labels
    r = 5
    ws.cell(row=r, column=1, value='').font = FONT_NOTE
    for i, yr in enumerate(years):
        c = i + 3
        if yr <= 2025:
            ws.cell(row=r, column=c, value='Historical').font = FONT_NOTE
        elif yr <= 2027:
            ws.cell(row=r, column=c, value='Phase 1').font = FONT_NOTE
        elif yr <= 2030:
            ws.cell(row=r, column=c, value='Phase 2').font = FONT_NOTE
        else:
            ws.cell(row=r, column=c, value='Phase 3').font = FONT_NOTE
        ws.cell(row=r, column=c).alignment = ALIGN_CENTER

    r = 7
    section_header(ws, r, 'Member Economics', len(yr_labels))
    r += 1

    row_items = [
        ('Beginning Members', 'M', [1.2, 2.0, None, None, None, None, None, None, None, None], FONT_INPUT),
        ('Gross Adds', 'M', [None]*10, FONT_INPUT),
        ('Churned Members', 'M', [None]*10, FONT_FORMULA),
        ('Ending Members', 'M', [2.0, 2.5, None, None, None, None, None, None, None, None], FONT_INPUT),
        ('Avg Members', 'M', [None]*10, FONT_FORMULA),
        ('Annual Churn Rate', '%', [None]*10, FONT_LINK),
        ('', '', [None]*10, None),
    ]

    for label, unit, vals, font in row_items:
        ws.cell(row=r, column=1, value=label).font = font or FONT_FORMULA
        ws.cell(row=r, column=2, value=unit).font = FONT_UNITS
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=i+3, value=v)
            cell.font = FONT_INPUT if v is not None else FONT_FORMULA
            if unit == 'M':
                cell.number_format = '0.0'
            elif unit == '%':
                cell.number_format = FMT_PCT
        r += 1

    section_header(ws, r, 'ARPU Build', len(yr_labels))
    r += 1

    arpu_items = [
        ('Subscription ARPU', '$/yr', [None, 262, None, None, None, None, None, None, None, None]),
        ('Non-Sub ARPU (Labs, Accessories)', '$/yr', [None, 40, None, None, None, None, None, None, None, None]),
        ('Blended ARPU', '$/yr', [None, 302, None, None, None, None, None, None, None, None]),
        ('Price Inflation (annual)', '%', [None]*10),
        ('', '', [None]*10),
    ]

    for label, unit, vals in arpu_items:
        ws.cell(row=r, column=1, value=label).font = FONT_FORMULA
        ws.cell(row=r, column=2, value=unit).font = FONT_UNITS
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=i+3, value=v)
            cell.font = FONT_INPUT if v is not None else FONT_FORMULA
            if unit == '$/yr':
                cell.number_format = FMT_DOLLARS
            elif unit == '%':
                cell.number_format = FMT_PCT
        r += 1

    section_header(ws, r, 'Revenue Summary', len(yr_labels))
    r += 1

    rev_items = [
        ('Subscription Revenue', '$M', [None, None, None, None, None, None, None, None, None, None]),
        ('Hardware Revenue', '$M', [None]*10),
        ('Non-Sub Revenue (Labs, Unite, Acc.)', '$M', [None]*10),
        ('Total Revenue', '$M', [200, 650, None, None, None, None, None, None, None, None]),
        ('YoY Revenue Growth', '%', [None]*10),
        ('', '', [None]*10),
        ('Bookings Run-Rate (reference)', '$M', [None, 1100, None, None, None, None, None, None, None, None]),
        ('Bookings-to-Revenue Ratio', 'x', [None, None, None, None, None, None, None, None, None, None]),
    ]

    for label, unit, vals in rev_items:
        ws.cell(row=r, column=1, value=label).font = FONT_FORMULA
        ws.cell(row=r, column=2, value=unit).font = FONT_UNITS
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=i+3, value=v)
            cell.font = FONT_INPUT if v is not None else FONT_FORMULA
            if unit == '$M':
                cell.number_format = FMT_MILLIONS
            elif unit == '%':
                cell.number_format = FMT_PCT
            elif unit == 'x':
                cell.number_format = '0.0x'
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4: P&L
# ═══════════════════════════════════════════════════════════════════════════════

def build_pnl(wb):
    ws = wb.create_sheet('P&L')
    ws.sheet_properties.tabColor = 'BF8F00'
    cols = {'A': 35, 'B': 5}
    for i in range(10):
        cols[get_column_letter(i+3)] = 14
    set_col_widths(ws, cols)

    ws.cell(row=1, column=1, value='WHOOP — Projected Income Statement').font = FONT_TITLE
    ws.cell(row=2, column=1, value='$M; Calendar Year; Scenario per Assumptions toggle').font = FONT_UNITS

    years = list(range(2024, 2034))
    yr_labels = [''] + ['Units'] + [str(y) for y in years]

    r = 4
    col_headers(ws, r, yr_labels)

    r = 6
    pnl_rows = [
        ('section', 'Revenue'),
        ('row', 'Total Revenue', '$M', FONT_LINK, '← From Revenue Build tab'),
        ('row', 'YoY Growth', '%', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Cost of Revenue'),
        ('row', 'Subscription COGS', '$M', FONT_FORMULA, ''),
        ('row', 'Hardware COGS', '$M', FONT_FORMULA, ''),
        ('row', 'Total COGS', '$M', FONT_FORMULA, ''),
        ('row', 'Gross Profit', '$M', FONT_FORMULA, ''),
        ('row', 'Gross Margin', '%', FONT_FORMULA, ''),
        ('row', '  Subscription GM', '%', FONT_LINK, '← From Assumptions C2'),
        ('row', '  Hardware GM', '%', FONT_LINK, '← From Assumptions C3'),
        ('blank',),
        ('section', 'Operating Expenses'),
        ('row', 'R&D', '$M', FONT_FORMULA, ''),
        ('row', '  R&D % Revenue', '%', FONT_LINK, '← From Assumptions C7/C8'),
        ('row', 'Sales & Marketing', '$M', FONT_FORMULA, ''),
        ('row', '  S&M % Revenue', '%', FONT_LINK, '← From Assumptions C9/C10'),
        ('row', 'General & Administrative', '$M', FONT_FORMULA, ''),
        ('row', '  G&A % Revenue', '%', FONT_LINK, '← From Assumptions C11/C12'),
        ('row', 'Stock-Based Compensation', '$M', FONT_FORMULA, ''),
        ('row', '  SBC % Revenue', '%', FONT_LINK, '← From Assumptions C13'),
        ('row', 'Total OpEx', '$M', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Profitability'),
        ('row', 'EBITDA', '$M', FONT_FORMULA, ''),
        ('row', 'EBITDA Margin', '%', FONT_FORMULA, ''),
        ('row', 'EBIT / Operating Income', '$M', FONT_FORMULA, ''),
        ('row', 'Operating Margin', '%', FONT_FORMULA, ''),
        ('row', 'Rule of 40 Score', '#', FONT_FORMULA, '= Growth % + FCF Margin %'),
        ('blank',),
        ('section', 'Below the Line'),
        ('row', 'Interest Income / (Expense)', '$M', FONT_INPUT, ''),
        ('row', 'Pre-Tax Income', '$M', FONT_FORMULA, ''),
        ('row', 'Taxes', '$M', FONT_FORMULA, ''),
        ('row', 'Net Income', '$M', FONT_FORMULA, ''),
        ('row', 'Net Margin', '%', FONT_FORMULA, ''),
    ]

    for item in pnl_rows:
        if item[0] == 'section':
            section_header(ws, r, item[1], len(yr_labels))
        elif item[0] == 'blank':
            pass
        elif item[0] == 'row':
            ws.cell(row=r, column=1, value=item[1]).font = FONT_FORMULA
            ws.cell(row=r, column=2, value=item[2]).font = FONT_UNITS
            if len(item) > 4 and item[4]:
                # Put note in last column
                ws.cell(row=r, column=13, value=item[4]).font = FONT_NOTE
            # Set font style for data cells
            for c in range(3, 13):
                ws.cell(row=r, column=c).font = item[3]
                if item[2] == '%':
                    ws.cell(row=r, column=c).number_format = FMT_PCT
                elif item[2] == '$M':
                    ws.cell(row=r, column=c).number_format = FMT_MILLIONS
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5: DCF
# ═══════════════════════════════════════════════════════════════════════════════

def build_dcf(wb):
    ws = wb.create_sheet('DCF')
    ws.sheet_properties.tabColor = 'C00000'
    cols = {'A': 35, 'B': 5}
    for i in range(10):
        cols[get_column_letter(i+3)] = 14
    set_col_widths(ws, cols)

    ws.cell(row=1, column=1, value='WHOOP — Discounted Cash Flow').font = FONT_TITLE
    ws.cell(row=2, column=1, value='$M; Calendar Year; BUILD YOUR FORMULAS HERE').font = Font(name='Calibri', size=11, color='CC0000', bold=True)

    years = list(range(2025, 2034))
    yr_labels = [''] + ['Units'] + [str(y) for y in years]

    r = 4
    col_headers(ws, r, yr_labels)

    r = 6
    dcf_rows = [
        ('section', 'Free Cash Flow Build'),
        ('row', 'EBIT', '$M', FONT_LINK, '← From P&L'),
        ('row', 'Less: Taxes on EBIT', '$M', FONT_FORMULA, ''),
        ('row', 'NOPAT', '$M', FONT_FORMULA, ''),
        ('row', 'Plus: D&A', '$M', FONT_FORMULA, ''),
        ('row', 'Less: CapEx', '$M', FONT_FORMULA, ''),
        ('row', 'Less: Change in NWC', '$M', FONT_FORMULA, ''),
        ('row', 'Less: SBC (real cost)', '$M', FONT_LINK, '← From P&L; SBC included per CLAUDE.md'),
        ('row', 'Unlevered Free Cash Flow', '$M', FONT_FORMULA, ''),
        ('row', 'FCF Margin', '%', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Discounting'),
        ('row', 'WACC', '%', FONT_LINK, '← From Assumptions D1-D5'),
        ('row', 'Discount Period', '#', FONT_FORMULA, 'Mid-year convention'),
        ('row', 'Discount Factor', 'x', FONT_FORMULA, ''),
        ('row', 'PV of FCF', '$M', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Terminal Value'),
        ('row', 'Terminal Year Revenue', '$M', FONT_LINK, ''),
        ('row', 'Terminal Year EBITDA', '$M', FONT_LINK, ''),
        ('row', 'Terminal Year FCF', '$M', FONT_LINK, ''),
        ('row', 'Exit Multiple (EV/Revenue)', 'x', FONT_LINK, '← From Assumptions D10'),
        ('row', 'Terminal Value (Exit Multiple)', '$M', FONT_FORMULA, ''),
        ('row', 'Terminal Value (Gordon Growth)', '$M', FONT_FORMULA, '= FCF × (1+g) / (WACC - g)'),
        ('row', 'TV Divergence Check', '%', FONT_FORMULA, 'If >25%, an assumption is inconsistent'),
        ('row', 'Selected Terminal Value', '$M', FONT_INPUT, ''),
        ('row', 'PV of Terminal Value', '$M', FONT_FORMULA, ''),
        ('row', 'TV as % of Total EV', '%', FONT_FORMULA, 'If >75%, explicit forecast doing too little'),
        ('blank',),
        ('section', 'Enterprise Value Summary'),
        ('row', 'Sum of PV of FCFs', '$M', FONT_FORMULA, ''),
        ('row', 'PV of Terminal Value', '$M', FONT_FORMULA, ''),
        ('row', 'Enterprise Value (DCF)', '$M', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Equity Bridge'),
        ('row', 'Enterprise Value', '$M', FONT_FORMULA, ''),
        ('row', 'Less: Debt', '$M', FONT_INPUT, '~$0 (minimal)'),
        ('row', 'Plus: Cash', '$M', FONT_INPUT, '$600-800M est. post-Series G'),
        ('row', 'Equity Value', '$M', FONT_FORMULA, ''),
        ('row', 'Fully Diluted Shares (M)', '#', FONT_INPUT, '~900M (PremierAlts est.)'),
        ('row', 'Equity Value per Share', '$/sh', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Real Options (Additive)'),
        ('row', 'Correlated Success EV', '$M', FONT_FORMULA, '← Probability × Payoff from Assumptions'),
        ('row', 'Mixed Outcomes EV', '$M', FONT_FORMULA, ''),
        ('row', 'Correlated Failure EV', '$M', FONT_FORMULA, ''),
        ('row', 'Expected Option Value', '$M', FONT_FORMULA, ''),
        ('row', 'Total EV (DCF + Options)', '$M', FONT_FORMULA, ''),
        ('blank',),
        ('section', 'Scenario Comparison'),
        ('row', 'DCF Base Case EV', '$M', FONT_FORMULA, ''),
        ('row', 'DCF + Options EV', '$M', FONT_FORMULA, ''),
        ('row', 'Series G Mark', '$M', FONT_INPUT, '10100'),
        ('row', 'Premium / (Discount) to Series G', '%', FONT_FORMULA, ''),
    ]

    for item in dcf_rows:
        if item[0] == 'section':
            section_header(ws, r, item[1], len(yr_labels))
        elif item[0] == 'blank':
            pass
        elif item[0] == 'row':
            ws.cell(row=r, column=1, value=item[1]).font = FONT_FORMULA
            ws.cell(row=r, column=2, value=item[2]).font = FONT_UNITS
            if len(item) > 4 and item[4]:
                ws.cell(row=r, column=12, value=item[4]).font = FONT_NOTE
            for c in range(3, 12):
                ws.cell(row=r, column=c).font = item[3]
                if item[2] == '%':
                    ws.cell(row=r, column=c).number_format = FMT_PCT
                elif item[2] == '$M':
                    ws.cell(row=r, column=c).number_format = FMT_MILLIONS
                elif item[2] == 'x':
                    ws.cell(row=r, column=c).number_format = '0.0x'
                elif item[2] == '$/sh':
                    ws.cell(row=r, column=c).number_format = FMT_SHARE_PRICE
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6: COMPS (FULLY POPULATED)
# ═══════════════════════════════════════════════════════════════════════════════

def build_comps(wb):
    ws = wb.create_sheet('Comps')
    ws.sheet_properties.tabColor = '7030A0'
    set_col_widths(ws, {
        'A': 22, 'B': 8, 'C': 10, 'D': 12, 'E': 12, 'F': 12, 'G': 12,
        'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12
    })

    ws.cell(row=1, column=1, value='WHOOP — Comparable Company Analysis').font = FONT_TITLE
    ws.cell(row=2, column=1, value='As of April 16, 2026; All USD').font = FONT_UNITS

    r = 4
    headers = ['Company', 'Ticker', 'Bucket', 'Mkt Cap ($B)', 'EV ($B)', 'LTM Rev ($B)',
               'EV/Rev LTM', 'EV/Rev NTM', 'EV/EBITDA', 'Gross Margin', 'Op Margin',
               'Rev Growth', 'Rule of 40', 'Lev. Beta']
    col_headers(ws, r, headers)

    r = 5
    ws.cell(row=r, column=1, value='').font = FONT_UNITS
    units = ['', '', '', '$B', '$B', '$B', 'x', 'x', 'x', '%', '%', '%', '#', '#']
    for i, u in enumerate(units):
        ws.cell(row=r, column=i+1, value=u).font = FONT_UNITS
        ws.cell(row=r, column=i+1).alignment = ALIGN_CENTER

    r = 7
    section_header(ws, r, 'Bucket 1: Consumer Hardware', 14)
    r += 1

    comps_data = [
        # B1
        ('Garmin', 'GRMN', 'B1', 51.2, 48.6, 7.25, 6.7, 6.2, 22.9, 0.587, 0.259, 0.15, 34, 0.97),
        # B2
        'SECTION:Bucket 2: Consumer Subscription Hybrid',
        ('Peloton', 'PTON', 'B2', 2.1, 2.8, 2.49, 1.2, 1.2, 16.4, 0.509, -0.014, -0.08, 5, 2.38),
        ('Spotify', 'SPOT', 'B2', 109.3, 102.6, 20.2, 5.1, 4.5, 38.8, 0.320, 0.128, 0.10, 27, 1.50),
        # B3
        'SECTION:Bucket 3: Health Data / Medical Device',
        ('Dexcom', 'DXCM', 'B3', 23.7, 23.1, 4.66, 5.0, 4.4, 19.8, 0.600, 0.200, 0.16, 39, 1.20),
        ('ResMed', 'RMD', 'B3', 33.2, 32.6, 5.40, 6.0, 5.5, 16.1, 0.594, 0.328, 0.10, 47, 0.93),
        ('Masimo', 'MASI', 'B3', 9.3, 9.7, 1.52, 6.4, 6.0, 28.2, 0.630, 0.275, 0.09, 37, 1.20),
        ('iRhythm', 'IRTC', 'B3', 4.0, 4.2, 0.75, 5.6, 4.8, 60.5, 0.706, -0.077, 0.26, 35, 1.26),
    ]

    for row_data in comps_data:
        if isinstance(row_data, str) and row_data.startswith('SECTION:'):
            r += 1
            section_header(ws, r, row_data.replace('SECTION:', ''), 14)
            r += 1
            continue

        fmts = [None, None, None, '0.0', '0.0', '0.00', '0.0x', '0.0x', '0.0x',
                FMT_PCT, FMT_PCT, FMT_PCT, '0', '0.00']

        for ci, val in enumerate(row_data):
            cell = ws.cell(row=r, column=ci+1, value=val)
            cell.font = FONT_INPUT
            if fmts[ci]:
                cell.number_format = fmts[ci]
            if ci >= 3:
                cell.alignment = ALIGN_RIGHT
        r += 1

    # Summary stats
    r += 2
    section_header(ws, r, 'Bucket Summary Statistics', 14)
    r += 1

    summary_headers = ['Bucket', '', '', '', '', '', 'EV/Rev LTM', 'EV/Rev NTM', '', 'Gross Margin', '', 'Rev Growth', 'Rule of 40', '']
    col_headers(ws, r, summary_headers)
    r += 1

    bucket_summary = [
        ('B1: Hardware', '', '', '', '', '', 6.7, 6.2, '', 0.587, '', 0.15, 34, ''),
        ('B2: Subscription (Peloton)', '', '', '', '', '', 1.2, 1.2, '', 0.509, '', -0.08, 5, ''),
        ('B2: Subscription (Spotify)', '', '', '', '', '', 5.1, 4.5, '', 0.320, '', 0.10, 27, ''),
        ('B3: Health Data (Median)', '', '', '', '', '', 5.8, 5.2, '', 0.615, '', 0.13, 38, ''),
        ('B3: Health Data (Range)', '', '', '', '', '', '5.0-6.4', '4.4-6.0', '', '', '', '', '', ''),
    ]

    for row_data in bucket_summary:
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=r, column=ci+1, value=val)
            cell.font = FONT_FORMULA
            if isinstance(val, float):
                if val < 1.5 and val > -0.5:
                    cell.number_format = FMT_PCT
                else:
                    cell.number_format = '0.0x'
            elif isinstance(val, int):
                cell.number_format = '0'
        r += 1

    # Bucket weighting calc
    r += 2
    section_header(ws, r, 'Implied WHOOP Valuation by Bucket Weighting', 14)
    r += 1

    ws.cell(row=r, column=1, value='Scenario').font = FONT_HEADER
    ws.cell(row=r, column=2, value='B1 Wt').font = FONT_HEADER
    ws.cell(row=r, column=3, value='B2 Wt').font = FONT_HEADER
    ws.cell(row=r, column=4, value='B3 Wt').font = FONT_HEADER
    ws.cell(row=r, column=5, value='B1 Mult').font = FONT_HEADER
    ws.cell(row=r, column=6, value='B2 Mult').font = FONT_HEADER
    ws.cell(row=r, column=7, value='B3 Mult').font = FONT_HEADER
    ws.cell(row=r, column=8, value='Blended').font = FONT_HEADER
    ws.cell(row=r, column=9, value='Impl. EV ($B)').font = FONT_HEADER
    ws.cell(row=r, column=10, value='vs. $10.1B').font = FONT_HEADER
    for c in range(1, 11):
        ws.cell(row=r, column=c).fill = FILL_HEADER
        ws.cell(row=r, column=c).alignment = ALIGN_CENTER
    r += 1

    scenarios = [
        ('Bear', 0.40, 0.40, 0.20, 6.7, 2.0, 5.8, None, None, None),
        ('Base', 0.20, 0.40, 0.40, 6.7, 3.5, 5.8, None, None, None),
        ('Bull', 0.10, 0.30, 0.60, 6.7, 4.0, 5.8, None, None, None),
        ('Series G Implied', 0.05, 0.25, 0.70, 6.7, 4.0, 5.8, None, None, None),
    ]

    for s in scenarios:
        for ci, val in enumerate(s):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_HEADER
            elif ci in (1, 2, 3):
                cell.font = FONT_INPUT
                cell.number_format = FMT_PCT
            elif ci in (4, 5, 6):
                cell.font = FONT_INPUT
                cell.number_format = '0.0x'
            elif ci == 7:
                cell.font = FONT_FORMULA
                cell.number_format = '0.0x'
            elif ci == 8:
                cell.font = FONT_FORMULA
                cell.number_format = '0.0'
            elif ci == 9:
                cell.font = FONT_FORMULA
                cell.number_format = FMT_PCT
        r += 1

    # Note
    r += 1
    ws.cell(row=r, column=1, value='Note: Blended = Σ(Wt × Mult). Implied EV = Blended × Revenue. Fill formulas.').font = FONT_NOTE
    r += 1
    ws.cell(row=r, column=1, value='Revenue base: $650M (2025 recognized) or $1,100M (bookings) — choose in Revenue Build tab.').font = FONT_NOTE


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 7: PRECEDENTS (FULLY POPULATED)
# ═══════════════════════════════════════════════════════════════════════════════

def build_precedents(wb):
    ws = wb.create_sheet('Precedents')
    ws.sheet_properties.tabColor = 'FF6600'
    set_col_widths(ws, {
        'A': 28, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 14, 'G': 10, 'H': 35
    })

    ws.cell(row=1, column=1, value='WHOOP — Precedent Transactions').font = FONT_TITLE
    ws.cell(row=2, column=1, value='2020-2026; $500M+ EV preferred; All USD').font = FONT_UNITS

    r = 4
    headers = ['Acquirer / Target', 'Date', 'EV ($B)', 'LTM Rev ($B)', 'EV/Rev', 'Regime', 'Bucket', 'Strategic Rationale']
    col_headers(ws, r, headers)

    r = 5
    deals = [
        ('Google / Fitbit', 'Jan 2021', 2.1, 1.43, 1.5, 'ZIRP-peak', 'B1', 'Wearable data acquisition; failed sub pivot'),
        ('Lululemon / Mirror', 'Jul 2020', 0.50, 0.17, 2.9, 'ZIRP-peak', 'B2', 'Connected fitness content; 89% write-off'),
        ('Teladoc / Livongo', 'Oct 2020', 18.5, 0.29, 63.8, 'ZIRP-peak', 'Cross', 'Digital health mega-deal; massive write-down'),
        ('Microsoft / Nuance', 'Mar 2022', 19.7, 1.48, 13.3, 'Transition', 'Cross', 'Healthcare AI voice/NLP platform'),
        ('Masimo / Sound United', 'Apr 2022', 1.03, 0.90, 1.1, 'Transition', 'B1', 'Consumer audio pivot; $675M value destruction'),
        ('CVS / Signify Health', 'Sep 2022', 8.0, 0.77, 10.4, 'Transition', 'B3', 'Home health + value-based care platform'),
        ('Amazon / One Medical', 'Feb 2023', 3.9, 1.1, 3.5, 'Transition', 'Cross', 'Primary care distribution; strategic access play'),
        ('J&J / Shockwave Medical', 'May 2024', 13.1, 0.73, 17.9, 'Current', 'B3', 'OUTLIER: IVL synergies drove 2x+ peer median'),
        ('BSci / Axonics', 'Nov 2024', 3.7, 0.37, 10.1, 'Current', 'B3', 'Neuromodulation bolt-on'),
        ('Stryker / Inari Medical', 'Feb 2025', 4.9, 0.60, 8.1, 'Current', 'B3', 'Venous thromboembolism devices'),
        ('Danaher / Masimo', 'Feb 2026', 9.9, 1.51, 6.6, 'Current', 'B3', 'Patient monitoring platform; pending close'),
    ]

    for d in deals:
        for ci, val in enumerate(d):
            cell = ws.cell(row=r, column=ci+1, value=val)
            cell.font = FONT_INPUT
            if ci == 2:  # EV
                cell.number_format = '0.0'
            elif ci == 3:  # Rev
                cell.number_format = '0.00'
            elif ci == 4:  # Multiple
                cell.number_format = '0.0x'
            elif ci == 7:  # Rationale
                cell.font = FONT_NOTE
                cell.alignment = ALIGN_LEFT
        r += 1

    r += 2
    section_header(ws, r, 'Summary by Regime & Bucket', 8)
    r += 1

    ws.cell(row=r, column=1, value='Current-regime B3 medtech: 6.6-10.1x (median ~8.5x)').font = FONT_FORMULA
    r += 1
    ws.cell(row=r, column=1, value='Connected fitness deals: uniformly value-destructive').font = FONT_FLAG
    r += 1
    ws.cell(row=r, column=1, value='Implied WHOOP range: $5-8B (control premium embedded; Series G is minority)').font = FONT_FORMULA


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 8: CAP TABLE
# ═══════════════════════════════════════════════════════════════════════════════

def build_cap_table(wb):
    ws = wb.create_sheet('Cap Table')
    ws.sheet_properties.tabColor = '00B050'
    set_col_widths(ws, {
        'A': 22, 'B': 14, 'C': 14, 'D': 16, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 25
    })

    ws.cell(row=1, column=1, value='WHOOP — Cap Table & Waterfall').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Sources: Pitchbook + Cap IQ; BUILD WATERFALL FORMULAS').font = FONT_UNITS

    r = 4
    headers = ['Round', 'Date', 'Amount ($M)', 'Post-Money ($B)', 'Price/Share ($)', 'Shares (M)', 'Ownership %', 'Pref Terms', 'Lead Investor(s)']
    col_headers(ws, r, headers)

    r = 5
    rounds = [
        ('Seed', '2012', 1.0, None, None, None, None, '1x non-part.', 'DraftKings founders'),
        ('Series A', 'Sep 2014', 4.0, None, None, None, None, '1x non-part.', 'Accomplice, Mousse Partners'),
        ('Series B', 'Sep 2016', 12.0, None, None, None, None, '1x non-part.', 'Foundry Group'),
        ('Series C', 'Oct 2018', 55.0, 0.43, None, None, None, '1x non-part.', 'IVP'),
        ('Series D', 'Oct 2020', 100.0, 1.20, None, None, None, '1x non-part.', 'IVP'),
        ('Series E', 'Aug 2021', 200.0, 2.40, None, None, None, '1x non-part.', 'SoftBank Vision Fund 2'),
        ('Series F', 'Aug 2021', 200.0, 3.60, 4.76, None, None, '1x non-part.', 'SoftBank Vision Fund 2'),
        ('Mezz/Bridge', '2022-2024', 46.0, None, None, None, None, 'Various', '5 interim rounds (VC + debt)'),
        ('Series G', 'Mar 2026', 575.0, 10.10, 11.23, None, 0.057, '1x non-part.', 'Collaborative Fund'),
    ]

    for rd in rounds:
        for ci, val in enumerate(rd):
            cell = ws.cell(row=r, column=ci+1, value=val)
            cell.font = FONT_INPUT
            if ci == 2:
                cell.number_format = '0.0'
            elif ci == 3 and val:
                cell.number_format = '0.00'
            elif ci == 4 and val:
                cell.number_format = FMT_SHARE_PRICE
            elif ci == 6 and val:
                cell.number_format = FMT_PCT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Total Equity Raised:').font = FONT_HEADER
    ws.cell(row=r, column=3, value=1030).font = FONT_FORMULA
    ws.cell(row=r, column=3).number_format = FMT_MILLIONS

    r += 2
    section_header(ws, r, 'Key Metrics', 9)
    r += 1

    metrics = [
        ('Authorized Preferred Shares', '537.9M', 'Pitchbook'),
        ('Outstanding Common Shares', '361.2M', 'Pitchbook'),
        ('Fully Diluted Shares (est.)', '~900M', 'PremierAlts/Forge'),
        ('Option Pool (est.)', '~10-15% of FD', 'Inferred from gap; FLAGGED_GAP'),
        ('All Preferences', '1x non-participating pari passu', 'Pitchbook'),
        ('Total Liquidation Preference', '~$1.03B (floor)', 'Sum of equity raised × 1x'),
    ]

    for label, val, source in metrics:
        ws.cell(row=r, column=1, value=label).font = FONT_FORMULA
        ws.cell(row=r, column=3, value=val).font = FONT_INPUT
        ws.cell(row=r, column=5, value=source).font = FONT_NOTE
        r += 1

    r += 2
    section_header(ws, r, 'Waterfall at Exit Scenarios (BUILD FORMULAS)', 9)
    r += 1

    headers2 = ['', '', 'Bear ($5B)', '', 'Base ($10B)', '', 'Bull ($20B)', '', '']
    col_headers(ws, r, headers2)
    r += 1

    waterfall_rows = [
        'Total Exit Proceeds',
        'Less: Liquidation Preferences',
        'Remaining for Pro-Rata',
        'Preferred Conversion Check (convert or take pref?)',
        'Common Share Proceeds',
        'Common Value per Share',
        'Preferred Value per Share',
        'Founder Equity Value (est. 12-18% FD)',
    ]

    for label in waterfall_rows:
        ws.cell(row=r, column=1, value=label).font = FONT_FORMULA
        for c in [3, 5, 7]:
            ws.cell(row=r, column=c).font = FONT_FORMULA
            ws.cell(row=r, column=c).number_format = FMT_MILLIONS
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 9: FOOTBALL FIELD (structure only)
# ═══════════════════════════════════════════════════════════════════════════════

def build_football_field(wb):
    ws = wb.create_sheet('Football Field')
    ws.sheet_properties.tabColor = '002060'
    set_col_widths(ws, {'A': 30, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14})

    ws.cell(row=1, column=1, value='WHOOP — Valuation Football Field').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Populated after DCF is complete').font = FONT_UNITS

    r = 4
    headers = ['Method', 'Low ($B)', 'Mid ($B)', 'High ($B)', 'Weight', 'Wtd Mid ($B)']
    col_headers(ws, r, headers)

    r = 5
    methods = [
        ('1. Intrinsic DCF (base case)', None, None, None, None, None),
        ('   + Real Options', None, None, None, None, None),
        ('2. Trading Comps (bucket-wtd)', 5.1, 5.6, 5.9, None, None),
        ('   + Growth Premium', None, None, None, None, None),
        ('3. Precedent Transactions', 5.0, 6.5, 8.0, None, None),
        ('4. Last-Round Implied', 10.1, 10.1, 10.1, None, None),
        ('5. Implied IPO Range', None, None, None, None, None),
        ('', None, None, None, None, None),
        ('Weighted Average', None, None, None, None, None),
        ('Series G Mark', 10.1, 10.1, 10.1, None, None),
        ('Premium / (Discount)', None, None, None, None, None),
    ]

    for m in methods:
        for ci, val in enumerate(m):
            cell = ws.cell(row=r, column=ci+1, value=val)
            if ci == 0:
                cell.font = FONT_FORMULA
            elif ci == 4:
                cell.font = FONT_INPUT
                cell.number_format = FMT_PCT
            else:
                cell.font = FONT_FORMULA if val is None else FONT_INPUT
                if isinstance(val, (int, float)):
                    cell.number_format = '0.0'
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 10: SENSITIVITY (structure only)
# ═══════════════════════════════════════════════════════════════════════════════

def build_sensitivity(wb):
    ws = wb.create_sheet('Sensitivity')
    ws.sheet_properties.tabColor = 'A5A5A5'
    set_col_widths(ws, {'A': 25, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14})

    ws.cell(row=1, column=1, value='WHOOP — Sensitivity Analysis').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Build data tables after DCF is complete').font = FONT_UNITS

    r = 4
    section_header(ws, r, 'Grid 1: WACC × Terminal Multiple → Enterprise Value ($B)', 8)
    r += 1
    ws.cell(row=r, column=1, value='WACC ↓ / Terminal EV/Rev →').font = FONT_HEADER
    term_mults = [3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0]
    for i, m in enumerate(term_mults):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = FONT_HEADER
        cell.number_format = '0.0x'
        cell.fill = FILL_HEADER
    r += 1
    waccs = [0.09, 0.095, 0.10, 0.105, 0.11, 0.115, 0.12]
    for w in waccs:
        cell = ws.cell(row=r, column=1, value=w)
        cell.font = FONT_HEADER
        cell.number_format = FMT_PCT
        cell.fill = FILL_HEADER
        r += 1

    r += 2
    section_header(ws, r, 'Grid 2: Revenue Growth × EBITDA Margin → Enterprise Value ($B)', 8)
    r += 1
    ws.cell(row=r, column=1, value='Growth ↓ / EBITDA Margin →').font = FONT_HEADER
    margins = [0.15, 0.20, 0.25, 0.28, 0.30, 0.32, 0.35]
    for i, m in enumerate(margins):
        cell = ws.cell(row=r, column=i+2, value=m)
        cell.font = FONT_HEADER
        cell.number_format = FMT_PCT
        cell.fill = FILL_HEADER
    r += 1
    growths = [0.50, 0.60, 0.70, 0.80, 0.90, 1.00, 1.10]
    for g in growths:
        cell = ws.cell(row=r, column=1, value=g)
        cell.font = FONT_HEADER
        cell.number_format = FMT_PCT
        cell.fill = FILL_HEADER
        r += 1

    r += 2
    section_header(ws, r, 'Grid 3: Churn Rate × ARPU → WHOOP-Specific Unit Economics', 8)
    r += 1
    ws.cell(row=r, column=1, value='Churn ↓ / ARPU ($/yr) →').font = FONT_HEADER
    arpus = [250, 275, 300, 325, 350, 375, 400]
    for i, a in enumerate(arpus):
        cell = ws.cell(row=r, column=i+2, value=a)
        cell.font = FONT_HEADER
        cell.number_format = FMT_DOLLARS
        cell.fill = FILL_HEADER
    r += 1
    churns = [0.10, 0.13, 0.15, 0.17, 0.20, 0.23, 0.25]
    for c in churns:
        cell = ws.cell(row=r, column=1, value=c)
        cell.font = FONT_HEADER
        cell.number_format = FMT_PCT
        cell.fill = FILL_HEADER
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 11: SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

def build_sources(wb):
    ws = wb.create_sheet('Sources')
    ws.sheet_properties.tabColor = '808080'
    set_col_widths(ws, {'A': 45, 'B': 60, 'C': 15})

    ws.cell(row=1, column=1, value='Research File Index').font = FONT_TITLE

    r = 3
    headers = ['File', 'Contents', 'Session']
    col_headers(ws, r, headers)

    r = 4
    sources = [
        ('research/whoop-research-brief.md', 'Structured research brief', 'S0'),
        ('research/whoop-pitchbook-data.md', 'Pitchbook: cap table, funding, financials', 'S1'),
        ('research/whoop-capiq-data.md', 'Cap IQ: dilution waterfall, investors', 'S1'),
        ('research/oura-pitchbook-data.md', 'Oura private comp profile ($11B)', 'S1'),
        ('research/peloton-10k-analysis.md', 'Peloton deep dive (revenue, margins, subs)', 'S1'),
        ('research/web-research-sweep.md', 'Market data, Forge, Garmin, Strava', 'S1'),
        ('research/revenue-bookings-reconciliation.md', '$1.1B bookings vs revenue, ARPU, churn grids', 'S1'),
        ('research/investor-deep-research.md', 'Series G investor analysis', 'S1'),
        ('research/stress-test-international-revenue.md', '60% intl claim verification', 'S1'),
        ('research/stress-test-revenue-reconciliation.md', 'Revenue stress test (base: $650M)', 'S1'),
        ('research/comp-dexcom-dxcm.md', 'Dexcom 10-K (5yr financials)', 'S2'),
        ('research/comp-garmin-grmn.md', 'Garmin 10-K (5yr, 5 segments)', 'S2'),
        ('research/comp-resmed-rmd.md', 'ResMed 10-K (connected device + SaaS)', 'S2'),
        ('research/comp-masimo-masi.md', 'Masimo 10-K (Danaher deal, Sound United)', 'S2'),
        ('research/comp-irhythm-irtc.md', 'iRhythm 10-K (cardiac wearable + SaaS)', 'S2'),
        ('research/comp-spotify-spot.md', 'Spotify 20-F (751M MAU, profitability inflection)', 'S2'),
        ('research/comp-peloton-pton.md', 'Peloton comp profile (trading multiples)', 'S2'),
        ('research/comp-trading-multiples-summary.md', 'Master multiples + bucket analysis', 'S2'),
        ('research/damodaran-wacc-inputs.md', 'ERP 4.23%, betas, risk-free 4.3%', 'S2'),
        ('research/comp-historical-multiples-v19.md', '3yr time variance: DXCM 13.8x→5.0x', 'S2'),
        ('research/peloton-10k-gaps.md', 'S&M/G&A split (XBRL), SBC history', 'S2'),
        ('research/secondary-market-pricing.md', 'Forge/Hiive/EquityZen, DLOM analysis', 'S2'),
        ('research/precedent-transactions.md', '14 deals, 3-bucket M&A analysis', 'S2'),
    ]

    for s in sources:
        for ci, val in enumerate(s):
            cell = ws.cell(row=r, column=ci+1, value=val)
            cell.font = FONT_FORMULA if ci == 0 else FONT_NOTE
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()

    build_cover(wb)
    build_assumptions(wb)
    build_revenue(wb)
    build_pnl(wb)
    build_dcf(wb)
    build_comps(wb)
    build_precedents(wb)
    build_cap_table(wb)
    build_football_field(wb)
    build_sensitivity(wb)
    build_sources(wb)

    # Freeze panes on key tabs
    for sheet_name in ['Assumptions', 'Revenue Build', 'P&L', 'DCF', 'Comps']:
        ws = wb[sheet_name]
        ws.freeze_panes = 'C5' if sheet_name != 'Assumptions' else 'C6'

    output_path = '/Users/landonprojects/whoop_val/model/whoop-master-model.xlsx'
    wb.save(output_path)
    print(f'Saved to {output_path}')
    print(f'Tabs: {[ws.title for ws in wb.worksheets]}')


if __name__ == '__main__':
    main()
