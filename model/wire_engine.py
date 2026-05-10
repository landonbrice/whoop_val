#!/usr/bin/env python3
"""Wire the unit-economics engine downstream of Assumptions.

Adds/updates:
  - Members tab (NEW): cohort engine 2024-2033 driving avg members
  - ARPU tab (NEW): single-line blended ARPU with annual inflation
  - Revenue Build tab (UPDATE): populate formulas linking Members × ARPU
  - FCF tab (NEW scaffold): architecture template
  - RealOptions tab (NEW scaffold): cluster framework
  - Scenarios tab (NEW scaffold): scenario summary template
  - Checks tab (NEW scaffold): diagnostic check rows

Cross-sheet links to Assumptions use col F (Active value) so scenario
selector at Assumptions!$B$3 flows through the entire model.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WB_PATH = 'model/whoop-master-model.xlsx'

# ── Style ──────────────────────────────────────────────────────────────────
F_INPUT    = Font(name='Calibri', size=10, color='0000FF')
F_FORMULA  = Font(name='Calibri', size=10, color='000000')
F_LINK     = Font(name='Calibri', size=10, color='008000')  # green = cross-sheet
F_LABEL    = Font(name='Calibri', size=10, color='000000')
F_NOTE     = Font(name='Calibri', size=9,  color='666666', italic=True)
F_HEADER   = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
F_SECTION  = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
F_TITLE    = Font(name='Calibri', size=14, color='000000', bold=True)
F_SUBTITLE = Font(name='Calibri', size=9,  color='666666', italic=True)
F_CHECK    = Font(name='Calibri', size=10, color='C00000', bold=True)
F_PHASE    = Font(name='Calibri', size=9,  color='666666', italic=True)

P_SECTION  = PatternFill('solid', start_color='305496')
P_HEADER   = PatternFill('solid', start_color='1F4E78')
P_SUB      = PatternFill('solid', start_color='D9E1F2')
P_DERIVED  = PatternFill('solid', start_color='F2F2F2')
P_DIAG     = PatternFill('solid', start_color='FFF2CC')
P_HIST     = PatternFill('solid', start_color='E7E6E6')

ALIGN_C    = Alignment(horizontal='center', vertical='center')
ALIGN_L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_R    = Alignment(horizontal='right',  vertical='center')

THIN       = Side(style='thin', color='D9D9D9')
BORDER_B   = Border(bottom=THIN)

# Year layout: A=Label, B=Units, C=2024..L=2033 (matches existing scaffolds)
YEARS = list(range(2024, 2034))   # 10 columns
YEAR_COL = {y: get_column_letter(3 + i) for i, y in enumerate(YEARS)}
PHASE = {2024: 'Historical', 2025: 'Historical',
         2026: 'Phase 1', 2027: 'Phase 1',
         2028: 'Phase 2', 2029: 'Phase 2', 2030: 'Phase 2',
         2031: 'Phase 3', 2032: 'Phase 3', 2033: 'Phase 3'}

# Assumptions row map (col F is Active value)
A = {  # ID -> row number on Assumptions tab
    'R3a': 8, 'R3b': 9, 'R3': 10, 'R4': 11, 'R1': 12, 'R2': 13,
    'R5': 14, 'R6': 15, 'R7': 16, 'R8': 17, 'R9': 18,
    'R10a': 19, 'R10b': 20, 'R10c': 21, 'R11': 22, 'R12': 23,
    'R14': 24, 'R15': 25, 'R16': 26, 'R18': 27, 'R19': 28,
    'C1': 31, 'C2': 32, 'C4': 33, 'C5': 34, 'C6': 35,
    'C7a': 36, 'C7b': 37, 'C7c': 38, 'C8a': 39, 'C8b': 40,
    'C9': 41, 'C9a': 42, 'C9b': 43, 'C9t': 44,
    'C10': 45, 'C10t': 46, 'C11': 47, 'C13': 48, 'C16': 49, 'C17': 50, 'C15': 51,
    'D1': 68, 'D8': 73, 'D11': 77, 'D13': 79,
    'V11a': 83, 'V11b': 84, 'V11c': 85,
    'I2': 90, 'I3': 91, 'I5': 93,
    'S1': 54, 'S10': 61, 'S11': 62, 'S12': 63,
}


def aref(id_):
    """Return cross-sheet reference to active value of an assumption."""
    return f"Assumptions!$F${A[id_]}"


def apply(cell, font=None, fill=None, fmt=None, align=None):
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align


def title_block(ws, title, subtitle):
    ws.merge_cells('A1:L1')
    ws['A1'].value = title
    apply(ws['A1'], F_TITLE, None, None, ALIGN_L)
    ws.row_dimensions[1].height = 22
    ws.merge_cells('A2:L2')
    ws['A2'].value = subtitle
    apply(ws['A2'], F_SUBTITLE, None, None, ALIGN_L)


def year_header(ws, header_row=4, phase_row=5):
    apply(ws.cell(row=header_row, column=2, value='Units'),
          F_HEADER, P_HEADER, None, ALIGN_C)
    for y in YEARS:
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=header_row, column=col, value=y),
              F_HEADER, P_HEADER, '0', ALIGN_C)
        apply(ws.cell(row=phase_row, column=col, value=PHASE[y]),
              F_PHASE, None, None, ALIGN_C)
        if y <= 2025:
            ws.cell(row=phase_row, column=col).fill = P_HIST
    ws.row_dimensions[header_row].height = 18


def section(ws, row, text, span=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, F_SECTION, P_SECTION, None,
          Alignment(horizontal='left', vertical='center', indent=1))
    ws.row_dimensions[row].height = 20


def label_cell(ws, row, label, units=None, indent=False):
    a = ws.cell(row=row, column=1, value=('  ' + label) if indent else label)
    apply(a, F_LABEL, None, None, ALIGN_L)
    if units is not None:
        apply(ws.cell(row=row, column=2, value=units), F_NOTE, None, None, ALIGN_C)


def col_widths(ws):
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 8
    for y in YEARS:
        ws.column_dimensions[YEAR_COL[y]].width = 11


# ── Members tab ────────────────────────────────────────────────────────────
def build_members(wb):
    if 'Members' in wb.sheetnames:
        del wb['Members']
    idx = wb.sheetnames.index('Assumptions') + 1
    ws = wb.create_sheet('Members', idx)
    col_widths(ws)
    ws.freeze_panes = 'C6'

    title_block(ws, 'WHOOP — Member Cohort Engine',
                'Members in millions; cohort engine drives revenue. '
                '2025 beginning treated as Y3+ legacy block (simplification — see note).')
    year_header(ws)

    # Track row numbers for cross-references
    R = {}
    row = 7

    # ─── Beginning members ───
    section(ws, row, '1. Beginning Members (start of year)'); row += 1
    R['beg_total'] = row
    label_cell(ws, row, 'Beginning Members (total)', 'M')
    # 2024 historical input
    apply(ws.cell(row=row, column=3, value=0.7),
          F_INPUT, P_HIST, '0.00', ALIGN_R)
    # 2025: link to Assumptions R3a (active)
    apply(ws.cell(row=row, column=4, value=f"={aref('R3a')}"),
          F_LINK, None, '0.00', ALIGN_R)
    # 2026+: = prior year ending
    for y in YEARS[2:]:
        c = YEAR_COL[y]; prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=ord(c)-64,
                      value=f"={prev}{row+9}"),  # row+9 = ending members row (placeholder updated after)
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    # Y1 cohort
    R['beg_y1'] = row
    label_cell(ws, row, 'of which: Year-1 cohort', 'M', indent=True)
    apply(ws.cell(row=row, column=3, value=0.5), F_INPUT, P_HIST, '0.00', ALIGN_R)
    apply(ws.cell(row=row, column=4, value=0), F_INPUT, None, '0.00', ALIGN_R)
    # 2026+: = prior year gross adds
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=ord(YEAR_COL[y])-64,
                      value=f"={prev}{row+11}"),  # gross adds row placeholder
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    # Y2 cohort
    R['beg_y2'] = row
    label_cell(ws, row, 'of which: Year-2 cohort', 'M', indent=True)
    apply(ws.cell(row=row, column=3, value=0.15), F_INPUT, P_HIST, '0.00', ALIGN_R)
    apply(ws.cell(row=row, column=4, value=0), F_INPUT, None, '0.00', ALIGN_R)
    # 2026+: = prior year Y1 survivors
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=ord(YEAR_COL[y])-64,
                      value=f"={prev}{row+5}"),  # Y1 survivors row placeholder
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    # Y3+ cohort
    R['beg_y3'] = row
    label_cell(ws, row, 'of which: Year-3+ cohort', 'M', indent=True)
    apply(ws.cell(row=row, column=3, value=0.05), F_INPUT, P_HIST, '0.00', ALIGN_R)
    # 2025 Y3+ = R3a (entire beginning treated as Y3+ legacy)
    apply(ws.cell(row=row, column=4, value=f"={aref('R3a')}"),
          F_LINK, None, '0.00', ALIGN_R)
    # 2026+: = prior year Y3+ end state (row 14 in spec terms)
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=ord(YEAR_COL[y])-64,
                      value=f"={prev}{row+5}"),  # Y3+ end state row placeholder
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1
    row += 1  # spacer

    # ─── Churn rates ───
    section(ws, row, '2. Churn Rates (cohort-aged, from Assumptions)'); row += 1

    R['churn_y1'] = row
    label_cell(ws, row, 'Year-1 churn rate', '%')
    for y in YEARS:
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=row, column=col, value=f"={aref('R10a')}"),
              F_LINK, None, '0.0%', ALIGN_R)
    row += 1

    R['churn_y2'] = row
    label_cell(ws, row, 'Year-2 churn rate', '%')
    for y in YEARS:
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=row, column=col, value=f"={aref('R10b')}"),
              F_LINK, None, '0.0%', ALIGN_R)
    row += 1

    R['churn_y3'] = row
    label_cell(ws, row, 'Year-3+ churn rate', '%')
    for y in YEARS:
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=row, column=col, value=f"={aref('R10c')}"),
              F_LINK, None, '0.0%', ALIGN_R)
    row += 1
    row += 1

    # ─── Churned members ───
    section(ws, row, '3. Churned Members'); row += 1

    R['churned_y1'] = row
    label_cell(ws, row, 'Y1 churned', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"={cl}{R['beg_y1']}*{cl}{R['churn_y1']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['churned_y2'] = row
    label_cell(ws, row, 'Y2 churned', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"={cl}{R['beg_y2']}*{cl}{R['churn_y2']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['churned_y3'] = row
    label_cell(ws, row, 'Y3+ churned', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"={cl}{R['beg_y3']}*{cl}{R['churn_y3']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['churned_total'] = row
    label_cell(ws, row, 'Total churned', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"=SUM({cl}{R['churned_y1']}:{cl}{R['churned_y3']})"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1
    row += 1

    # ─── Surviving members (end state by cohort) ───
    section(ws, row, '4. Surviving Members (end of year, before new adds)'); row += 1

    R['surv_y1'] = row
    label_cell(ws, row, 'Y1 survivors', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['beg_y1']}-{cl}{R['churned_y1']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['surv_y2'] = row
    label_cell(ws, row, 'Y2 survivors', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['beg_y2']}-{cl}{R['churned_y2']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['surv_y3'] = row
    label_cell(ws, row, 'Y3+ end state (incl. Y2 graduates)', 'M', indent=True)
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        # Y3+ end = Y3+ begin - Y3+ churned + Y2 survivors (graduating)
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['beg_y3']}-{cl}{R['churned_y3']}+{cl}{R['surv_y2']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1
    row += 1

    # ─── Gross adds ───
    section(ws, row, '5. New Members (Gross Adds — input per year)'); row += 1

    R['gross_adds'] = row
    label_cell(ws, row, 'Gross Adds', 'M')
    # 2024 historical: input
    apply(ws.cell(row=row, column=3, value=0.6), F_INPUT, P_HIST, '0.00', ALIGN_R)
    # 2025: derived to back-solve to R3b ending: gross adds = R3b - Y3+_end_state
    apply(ws.cell(row=row, column=4,
                  value=f"={aref('R3b')}-D{R['surv_y3']}"),
          F_FORMULA, P_DERIVED, '0.00', ALIGN_R)
    # 2026: link to Assumptions R18
    apply(ws.cell(row=row, column=5, value=f"={aref('R18')}"),
          F_LINK, None, '0.00', ALIGN_R)
    # 2027-2033: input priors per build-spec (decelerating gross adds)
    priors = {2027: 1.7, 2028: 1.4, 2029: 1.2, 2030: 1.1, 2031: 1.0, 2032: 0.95, 2033: 0.9}
    for y, val in priors.items():
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=row, column=col, value=val), F_INPUT, None, '0.00', ALIGN_R)
    row += 1
    row += 1

    # ─── Ending members ───
    section(ws, row, '6. Ending & Average Members'); row += 1

    R['ending'] = row
    label_cell(ws, row, 'Ending Members', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        # End = Y1 surv + Y3+ end state + Gross adds (Y2 surv folded into Y3+ end state)
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['surv_y1']}+{cl}{R['surv_y3']}+{cl}{R['gross_adds']}"),
              F_FORMULA, None, '0.00', ALIGN_R)
    row += 1

    R['avg'] = row
    label_cell(ws, row, 'Average Members', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"=({cl}{R['beg_total']}+{cl}{R['ending']})/2"),
              F_FORMULA, P_DERIVED, '0.00', ALIGN_R)
    row += 1
    row += 1

    # Now we need to fix forward-references in beg_total/beg_y1/beg_y2/beg_y3 that
    # used row+9, row+11, row+5 placeholders. Update them now that we know exact rows.
    # beg_total 2026+ = prior year ending
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]; col = 3 + YEARS.index(y)
        ws.cell(row=R['beg_total'], column=col).value = f"={prev}{R['ending']}"
    # beg_y1 2026+ = prior year gross adds
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]; col = 3 + YEARS.index(y)
        ws.cell(row=R['beg_y1'], column=col).value = f"={prev}{R['gross_adds']}"
    # beg_y2 2026+ = prior year Y1 survivors
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]; col = 3 + YEARS.index(y)
        ws.cell(row=R['beg_y2'], column=col).value = f"={prev}{R['surv_y1']}"
    # beg_y3 2026+ = prior year Y3+ end state
    for y in YEARS[2:]:
        prev = YEAR_COL[y - 1]; col = 3 + YEARS.index(y)
        ws.cell(row=R['beg_y3'], column=col).value = f"={prev}{R['surv_y3']}"

    # ─── Diagnostics ───
    section(ws, row, '7. Diagnostics (output, not input)'); row += 1

    label_cell(ws, row, 'Net Adds (Gross - Churned)', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['gross_adds']}-{cl}{R['churned_total']}"),
              F_FORMULA, P_DIAG, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Net Member Growth %', '%')
    for y in YEARS[1:]:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]; prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['ending']}/{prev}{R['ending']}-1"),
              F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Blended Annual Churn Rate (output)', '%')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"=IFERROR({cl}{R['churned_total']}/{cl}{R['beg_total']},0)"),
              F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1

    label_cell(ws, row, '2025 ending vs Assumptions R3b (consistency)', '%')
    apply(ws.cell(row=row, column=4,
                  value=f"=D{R['ending']}/{aref('R3b')}-1"),
          F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1
    row += 1

    # Note
    note_text = ('NOTE: 2025 beginning is treated as Y3+ legacy block (simplification). '
                 'Replace with audited cohort splits when S-1 data is available. '
                 '2024 column is historical reference; engine fully runs 2025 onwards.')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    apply(ws.cell(row=row, column=1, value=note_text), F_NOTE, None, None, ALIGN_L)
    row += 1

    return R


# ── ARPU tab ───────────────────────────────────────────────────────────────
def build_arpu(wb):
    if 'ARPU' in wb.sheetnames:
        del wb['ARPU']
    idx = wb.sheetnames.index('Members') + 1
    ws = wb.create_sheet('ARPU', idx)
    col_widths(ws)
    ws.freeze_panes = 'C6'

    title_block(ws, 'WHOOP — ARPU Build',
                'Single-line blended ARPU with annual price inflation. '
                'Decompose into 5 components (sub/tier/accessories/Labs/BP) when sourcing improves.')
    year_header(ws)

    R = {}
    row = 7

    section(ws, row, '1. Blended ARPU'); row += 1
    R['arpu'] = row
    label_cell(ws, row, 'Blended ARPU ($/yr)', '$')
    # 2024 historical: input
    apply(ws.cell(row=row, column=3, value=280), F_INPUT, P_HIST, '$#,##0', ALIGN_R)
    # 2025: link to Assumptions R4
    apply(ws.cell(row=row, column=4, value=f"={aref('R4')}"),
          F_LINK, None, '$#,##0', ALIGN_R)
    # 2026+: prior year × (1 + inflation)
    for y in YEARS[2:]:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]; prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=col,
                      value=f"={prev}{row}*(1+{aref('R16')})"),
              F_FORMULA, None, '$#,##0', ALIGN_R)
    row += 1

    label_cell(ws, row, 'YoY ARPU growth', '%')
    for y in YEARS[1:]:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]; prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{R['arpu']}/{prev}{R['arpu']}-1"),
              F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1
    row += 1

    section(ws, row, '2. Future Decomposition (placeholder)'); row += 1
    decomp_lines = [
        ('Base subscription ($)',    'Per methodology Sec 3'),
        ('Tier premium ($)',          'Peak/Life upgrade above Core'),
        ('Accessories/straps ($)',    'Replacement cycle'),
        ('Advanced Labs ($)',         'Attach × test price (R14 × R15) — bucket-3 thesis'),
        ('BP / Healthspan ($)',       'Attach × price'),
    ]
    for label, note in decomp_lines:
        label_cell(ws, row, label, '$', indent=True)
        apply(ws.cell(row=row, column=12, value=note), F_NOTE, None, None, ALIGN_L)
        row += 1
    row += 1

    note_text = ('NOTE: Decomposition deferred per build directive. R4 base $338 is single-line '
                 'placeholder; Labs attach (R14) and price (R15) flagged as bucket-3 thesis variables.')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    apply(ws.cell(row=row, column=1, value=note_text), F_NOTE, None, None, ALIGN_L)

    return R


# ── Revenue Build tab (UPDATE existing) ────────────────────────────────────
def build_revenue(wb, members_R, arpu_R):
    ws = wb['Revenue Build']

    # Unmerge any data-row merged cells FIRST (else cell.value assignment fails)
    merged = list(ws.merged_cells.ranges)
    for mr in merged:
        if mr.min_row >= 6:
            ws.unmerge_cells(str(mr))

    # Wipe data rows (keep title/header rows 1-5)
    for row in range(6, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill('none')

    col_widths(ws)
    ws.freeze_panes = 'C6'

    # Re-paint header row coloring (matches new style)
    apply(ws.cell(row=4, column=2), F_HEADER, P_HEADER, None, ALIGN_C)
    for y in YEARS:
        col = 3 + YEARS.index(y)
        apply(ws.cell(row=4, column=col), F_HEADER, P_HEADER, '0', ALIGN_C)
        apply(ws.cell(row=5, column=col), F_PHASE, None, None, ALIGN_C)
        if y <= 2025:
            ws.cell(row=5, column=col).fill = P_HIST

    row = 7
    section(ws, row, '1. Member Economics (linked from Members tab)'); row += 1

    label_cell(ws, row, 'Beginning Members', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=Members!{cl}{members_R['beg_total']}"),
              F_LINK, None, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Gross Adds', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=Members!{cl}{members_R['gross_adds']}"),
              F_LINK, None, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Churned Members', 'M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=Members!{cl}{members_R['churned_total']}"),
              F_LINK, None, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Ending Members', 'M')
    end_row = row
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=Members!{cl}{members_R['ending']}"),
              F_LINK, None, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Avg Members', 'M')
    avg_row = row
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=Members!{cl}{members_R['avg']}"),
              F_LINK, None, '0.00', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Blended Annual Churn (output)', '%')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"=IFERROR({cl}{end_row-3}/{cl}{end_row-3-99}, 0)"
                      if False else  # not used
                      f"=IFERROR(Members!{cl}{members_R['churned_total']}/Members!{cl}{members_R['beg_total']}, 0)"),
              F_LINK, P_DIAG, '0.0%', ALIGN_R)
    row += 1
    row += 1

    section(ws, row, '2. ARPU (linked from ARPU tab)'); row += 1
    label_cell(ws, row, 'Blended ARPU', '$/yr')
    arpu_row_link = row
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col, value=f"=ARPU!{cl}{arpu_R['arpu']}"),
              F_LINK, None, '$#,##0', ALIGN_R)
    row += 1
    row += 1

    section(ws, row, '3. Revenue Build'); row += 1

    label_cell(ws, row, 'Recognized Revenue (Avg × ARPU)', '$M')
    rev_row = row
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{avg_row}*{cl}{arpu_row_link}"),
              F_FORMULA, P_DERIVED, '$#,##0', ALIGN_R)
    row += 1

    # Deferred revenue is a BALANCE SHEET effect (Δ in deferred rev flows through NWC on FCF tab),
    # not an adjustment to recognized revenue. Avg × ARPU is itself the recognized number per ASC 606.
    label_cell(ws, row, 'Memo: Δ Deferred Rev Balance (flows to NWC, not P&L)', '$M')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{rev_row}*{aref('R19')}"),
              F_FORMULA, P_DIAG, '$#,##0;($#,##0)', ALIGN_R)
    row += 1

    label_cell(ws, row, 'YoY Revenue Growth', '%')
    for y in YEARS[1:]:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]; prev = YEAR_COL[y - 1]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{rev_row}/{prev}{rev_row}-1"),
              F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1
    row += 1

    section(ws, row, '4. Bookings Reconciliation'); row += 1

    label_cell(ws, row, 'Implied Bookings (End × ARPU)', '$M')
    book_row = row
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"={cl}{end_row}*{cl}{arpu_row_link}"),
              F_FORMULA, None, '$#,##0', ALIGN_R)
    row += 1

    label_cell(ws, row, 'Bookings/Revenue Ratio', 'x')
    for y in YEARS:
        col = 3 + YEARS.index(y); cl = YEAR_COL[y]
        apply(ws.cell(row=row, column=col,
                      value=f"=IFERROR({cl}{book_row}/{cl}{rev_row}, 0)"),
              F_FORMULA, P_DIAG, '0.00"x"', ALIGN_R)
    row += 1

    label_cell(ws, row, '2025 Recognized vs Assumptions R1 (consistency)', '%')
    apply(ws.cell(row=row, column=4,
                  value=f"=D{rev_row}/{aref('R1')}-1"),
          F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1


# ── Scaffold tabs ──────────────────────────────────────────────────────────
def build_scaffold(wb, name, after_sheet, sections, subtitle=''):
    if name in wb.sheetnames:
        del wb[name]
    idx = wb.sheetnames.index(after_sheet) + 1
    ws = wb.create_sheet(name, idx)
    col_widths(ws)
    ws.freeze_panes = 'C6'

    title_block(ws, f'WHOOP — {name}', subtitle)
    year_header(ws)

    row = 7
    for sec_title, lines in sections:
        section(ws, row, sec_title); row += 1
        for label, units, note in lines:
            label_cell(ws, row, label, units)
            if note:
                apply(ws.cell(row=row, column=12, value=note), F_NOTE, None, None, ALIGN_L)
            row += 1
        row += 1


def build_fcf(wb):
    sections = [
        ('1. EBIT to NOPAT', [
            ('EBIT', '$M', 'Link from P&L'),
            ('Less: Cash taxes on EBIT', '$M', 'EBIT × tax rate (Assumptions C15)'),
            ('NOPAT', '$M', 'EBIT - taxes'),
        ]),
        ('2. NOPAT to Unlevered FCF', [
            ('Plus: D&A', '$M', 'Link from P&L'),
            ('Less: CapEx', '$M', 'Revenue × C13'),
            ('Less: Change in NWC', '$M', 'Δ(deferred rev) - Δ(inventory) - Δ(other WC)'),
            ('Note: SBC is REAL COST in P&L — do NOT add back', '', 'Per methodology Sec 6'),
            ('Unlevered Free Cash Flow', '$M', 'Sum'),
        ]),
        ('3. FCF Diagnostics', [
            ('FCF Margin %', '%', 'FCF / Revenue'),
            ('Rule of 40 (Growth + FCF Margin)', '%', 'Should be ≥40 in Phase 1-2'),
        ]),
    ]
    build_scaffold(wb, 'FCF', 'P&L', sections,
                   'Free cash flow architecture; populate after P&L is wired')


def build_real_options(wb):
    if 'RealOptions' in wb.sheetnames:
        del wb['RealOptions']
    idx = wb.sheetnames.index('DCF') + 1
    ws = wb.create_sheet('RealOptions', idx)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    title_block(ws, 'WHOOP — Real Options',
                'Cluster framework: correlated outcomes, additive to DCF base case. NEVER embedded in DCF.')

    # Headers
    headers = ['Cluster', 'Description', 'Value ($B)', 'Probability', 'Weighted Value ($B)']
    for i, h in enumerate(headers, 1):
        apply(ws.cell(row=4, column=i, value=h), F_HEADER, P_HEADER, None, ALIGN_C)

    row = 5
    clusters = [
        ('Correlated success',
         'Healthcare reimbursement + B2B clinical + Women\'s health all execute',
         5.5, 0.18),
        ('Mixed outcomes',
         'Geographic expansion works; healthcare stalls partially',
         2.5, 0.45),
        ('Correlated failure',
         'Healthcare thesis stalls; related options fail together',
         0.75, 0.37),
    ]
    for cluster, desc, val, prob in clusters:
        apply(ws.cell(row=row, column=1, value=cluster), F_LABEL, None, None, ALIGN_L)
        apply(ws.cell(row=row, column=2, value=desc), F_NOTE, None, None, ALIGN_L)
        apply(ws.cell(row=row, column=3, value=val), F_INPUT, None, '$#,##0.00', ALIGN_R)
        apply(ws.cell(row=row, column=4, value=prob), F_INPUT, None, '0.0%', ALIGN_R)
        apply(ws.cell(row=row, column=5, value=f"=C{row}*D{row}"), F_FORMULA, P_DERIVED, '$#,##0.00', ALIGN_R)
        row += 1

    apply(ws.cell(row=row, column=1, value='Probability sum (must = 100%)'),
          F_LABEL, None, None, ALIGN_L)
    apply(ws.cell(row=row, column=4, value=f"=SUM(D5:D7)"),
          F_FORMULA, P_DIAG, '0.0%', ALIGN_R)
    row += 1
    apply(ws.cell(row=row, column=1, value='Total Real Option Value (additive to DCF)'),
          F_LABEL, None, None, ALIGN_L)
    apply(ws.cell(row=row, column=5, value=f"=SUM(E5:E7)"),
          F_FORMULA, P_DERIVED, '$#,##0.00', ALIGN_R)
    row += 2

    # Individual options (probability lookup, not summed)
    section(ws, row, 'Individual Option Probabilities (reference, not additive)', span=5); row += 1
    individuals = [
        ('Healthcare reimbursement', '20-30% — upgraded by CMS Innovation Center selection (Apr 2026)'),
        ('International expansion', '40-60% — note ~60% of sales already intl, partially in base'),
        ('B2B enterprise / clinical', '25-40% — Unite already ~20% of recurring, partially in base'),
        ('Women\'s health / fertility', '30-50% — 150% YoY women membership growth'),
        ('Noninvasive glucose monitoring', '3-7% — deliberately conservative; hardest unsolved problem'),
    ]
    for opt, note in individuals:
        apply(ws.cell(row=row, column=1, value=opt), F_LABEL, None, None, ALIGN_L)
        apply(ws.cell(row=row, column=2, value=note), F_NOTE, None, None, ALIGN_L)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1


def build_scenarios(wb):
    if 'Scenarios' in wb.sheetnames:
        del wb['Scenarios']
    idx = wb.sheetnames.index('Football Field') + 1
    ws = wb.create_sheet('Scenarios', idx)

    ws.column_dimensions['A'].width = 36
    for col_letter in ['B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions['E'].width = 50

    title_block(ws, 'WHOOP — Scenario Summary',
                'Scenario EV by method; weight sensitivity grid. Populate after DCF + Comps are wired.')

    headers = ['Method', 'Bear ($B)', 'Base ($B)', 'Bull ($B)', 'Notes']
    for i, h in enumerate(headers, 1):
        apply(ws.cell(row=4, column=i, value=h), F_HEADER, P_HEADER, None, ALIGN_C)

    methods = [
        ('1. Intrinsic DCF (base)', 'Link from DCF tab'),
        ('   + Real Options', 'Link from RealOptions tab'),
        ('2. Trading Comps (3-bucket)', 'Link from Comps tab'),
        ('3. Precedent Transactions', 'Link from Precedents tab'),
        ('4. Last-Round Implied', 'Series G mark = $10.1B (constant)'),
        ('5. Implied IPO Trading Range', 'Link from Assumptions I5'),
    ]
    row = 5
    for m, note in methods:
        apply(ws.cell(row=row, column=1, value=m), F_LABEL, None, None, ALIGN_L)
        for col in [2, 3, 4]:
            apply(ws.cell(row=row, column=col), F_FORMULA, None, '$#,##0.00', ALIGN_R)
        apply(ws.cell(row=row, column=5, value=note), F_NOTE, None, None, ALIGN_L)
        row += 1
    row += 1

    section(ws, row, 'Probability Weight Sensitivity', span=5); row += 1
    weight_schemes = [
        ('Pessimistic', 0.35, 0.50, 0.15),
        ('Neutral (base)', 0.25, 0.50, 0.25),
        ('Optimistic', 0.15, 0.55, 0.30),
    ]
    apply(ws.cell(row=row, column=1, value='Weight Scheme'), F_HEADER, P_HEADER, None, ALIGN_C)
    apply(ws.cell(row=row, column=2, value='Bear Wt'), F_HEADER, P_HEADER, None, ALIGN_C)
    apply(ws.cell(row=row, column=3, value='Base Wt'), F_HEADER, P_HEADER, None, ALIGN_C)
    apply(ws.cell(row=row, column=4, value='Bull Wt'), F_HEADER, P_HEADER, None, ALIGN_C)
    apply(ws.cell(row=row, column=5, value='Weighted EV ($B)'), F_HEADER, P_HEADER, None, ALIGN_C)
    row += 1
    for name, bw, base_w, bullw in weight_schemes:
        apply(ws.cell(row=row, column=1, value=name), F_LABEL, None, None, ALIGN_L)
        apply(ws.cell(row=row, column=2, value=bw), F_INPUT, None, '0.0%', ALIGN_R)
        apply(ws.cell(row=row, column=3, value=base_w), F_INPUT, None, '0.0%', ALIGN_R)
        apply(ws.cell(row=row, column=4, value=bullw), F_INPUT, None, '0.0%', ALIGN_R)
        apply(ws.cell(row=row, column=5), F_FORMULA, P_DERIVED, '$#,##0.00', ALIGN_R)
        row += 1


def build_checks(wb):
    if 'Checks' in wb.sheetnames:
        del wb['Checks']
    idx = wb.sheetnames.index('Sensitivity') + 1
    ws = wb.create_sheet('Checks', idx)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50

    title_block(ws, 'WHOOP — Diagnostic Checks',
                'Automated integrity checks. ANY red FAIL must be resolved before output is trusted.')

    headers = ['Check', 'Computed', 'Threshold', 'Status', 'Notes']
    for i, h in enumerate(headers, 1):
        apply(ws.cell(row=4, column=i, value=h), F_HEADER, P_HEADER, None, ALIGN_C)

    checks = [
        ('Revenue: Avg Members × ARPU ≈ R1', '=\'Revenue Build\'!D17/Assumptions!$F$12-1', '±5%',
         'PASS if |delta| < 5%'),
        ('Members 2025 ending = R3b', '=Members!D24/Assumptions!$F$9-1', '±2%',
         'PASS if |delta| < 2%'),
        ('Bucket weights sum to 100%', '=Assumptions!$F$86', '100%', 'V11_chk row'),
        ('TV < 75% of EV (DCF)', '', '<75%', 'Build after DCF wired'),
        ('Exit multiple ↔ Gordon Growth divergence', '', '<25%', 'Build after DCF wired'),
        ('S&M ≥ Paid CAC + Brand', '', '>0', 'Build after P&L wired'),
        ('Real options probability sum', '=RealOptions!D8', '100%',
         'PASS if exactly 100%'),
        ('WACC > Perpetuity growth', '=Assumptions!$F$73-Assumptions!$F$79', '>0',
         'Required for Gordon Growth'),
        ('Cap structure: shares > 0', '=Assumptions!$F$54', '>0', 'FD shares populated'),
        ('Rule of 40 by Phase 3', '', '≥20', 'Build after FCF wired'),
    ]
    row = 5
    for check, formula, threshold, note in checks:
        apply(ws.cell(row=row, column=1, value=check), F_LABEL, None, None, ALIGN_L)
        if formula:
            c = ws.cell(row=row, column=2, value=formula)
            apply(c, F_FORMULA, P_DERIVED, '0.00%' if 'delta' in note or '%' in threshold else '0.00', ALIGN_R)
        apply(ws.cell(row=row, column=3, value=threshold), F_NOTE, None, None, ALIGN_C)
        apply(ws.cell(row=row, column=4, value='—'), F_CHECK, None, None, ALIGN_C)
        apply(ws.cell(row=row, column=5, value=note), F_NOTE, None, None, ALIGN_L)
        row += 1


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    wb = load_workbook(WB_PATH)

    members_R = build_members(wb)
    arpu_R = build_arpu(wb)
    build_revenue(wb, members_R, arpu_R)
    build_fcf(wb)
    build_real_options(wb)
    build_scenarios(wb)
    build_checks(wb)

    wb.save(WB_PATH)
    print(f'Wrote {WB_PATH}')
    print(f'Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
