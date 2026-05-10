#!/usr/bin/env python3
"""Wire the Assumptions tab from research base.

Restructures /model/whoop-master-model.xlsx Assumptions sheet only.
- Adds scenario selector (B3) + Active value column (col F) wired across rows
- Splits R3 → R3a/R3/R3b (begin/avg/end), R10 → R10a/R10b/R10c (cohort churn)
- Makes R1 (recognized revenue) DERIVED = avg members × blended ARPU
- Adds missing critical assumptions (cap structure, gross adds, Labs attach,
  precedent transactions); FLAGGED_GAP yellow highlight on unsourced criticals
- Removes Real Options from this tab (lives on RealOptions tab, future)
- IB conventions: blue inputs, black formulas, green cross-sheet links
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WB_PATH = 'model/whoop-master-model.xlsx'

# ── Style ──────────────────────────────────────────────────────────────────
F_INPUT    = Font(name='Calibri', size=10, color='0000FF')
F_FORMULA  = Font(name='Calibri', size=10, color='000000')
F_LINK     = Font(name='Calibri', size=10, color='008000')  # reserved for cross-sheet
F_LABEL    = Font(name='Calibri', size=10, color='000000')
F_NOTE     = Font(name='Calibri', size=9,  color='666666', italic=True)
F_TIER     = Font(name='Calibri', size=9,  color='666666')
F_FLAG     = Font(name='Calibri', size=10, color='C00000', bold=True)
F_HEADER   = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
F_SECTION  = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
F_TITLE    = Font(name='Calibri', size=14, color='000000', bold=True)
F_TOGGLE   = Font(name='Calibri', size=11, color='0000FF', bold=True)

P_TITLE    = PatternFill('solid', start_color='FFFFFF')
P_SECTION  = PatternFill('solid', start_color='305496')
P_HEADER   = PatternFill('solid', start_color='1F4E78')
P_FLAG     = PatternFill('solid', start_color='FFF2CC')
P_DERIVED  = PatternFill('solid', start_color='F2F2F2')
P_TOGGLE   = PatternFill('solid', start_color='DDEBF7')

ALIGN_C    = Alignment(horizontal='center', vertical='center')
ALIGN_L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_R    = Alignment(horizontal='right',  vertical='center')

THIN       = Side(style='thin', color='D9D9D9')
BORDER_B   = Border(bottom=THIN)


# ── Helpers ────────────────────────────────────────────────────────────────
def active_formula(row: int) -> str:
    return f'=IF($B$3="Bear",C{row},IF($B$3="Bull",E{row},D{row}))'


def section_header(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.fill = P_SECTION
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22


def write_input(ws, row, id_, desc, bear, base, bull, tier, status, notes,
                fmt='General', flagged=False):
    ws.cell(row=row, column=1, value=id_).font = F_LABEL
    desc_cell = ws.cell(row=row, column=2, value=desc); desc_cell.font = F_LABEL
    desc_cell.alignment = ALIGN_L
    for col, val in zip([3, 4, 5], [bear, base, bull]):
        c = ws.cell(row=row, column=col, value=val)
        c.font = F_INPUT
        c.number_format = fmt
        c.alignment = ALIGN_R
    fc = ws.cell(row=row, column=6, value=active_formula(row))
    fc.font = F_FORMULA
    fc.number_format = fmt
    fc.alignment = ALIGN_R
    ws.cell(row=row, column=7, value=tier).font = F_TIER
    sc = ws.cell(row=row, column=8, value=status)
    sc.font = F_FLAG if flagged else F_TIER
    nc = ws.cell(row=row, column=9, value=notes); nc.font = F_NOTE
    nc.alignment = ALIGN_L
    if flagged:
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = P_FLAG


def write_derived(ws, row, id_, desc, formulas, tier, status, notes, fmt='General'):
    """formulas: dict with optional keys 'C','D','E'; if missing, blank in that scen.
       Active (F) always uses the active_formula picker."""
    ws.cell(row=row, column=1, value=id_).font = F_LABEL
    desc_cell = ws.cell(row=row, column=2, value=desc); desc_cell.font = F_LABEL
    desc_cell.alignment = ALIGN_L
    for col, letter in zip([3, 4, 5], ['C', 'D', 'E']):
        if letter in formulas:
            c = ws.cell(row=row, column=col, value=formulas[letter])
            c.font = F_FORMULA
            c.fill = P_DERIVED
            c.number_format = fmt
            c.alignment = ALIGN_R
    fc = ws.cell(row=row, column=6, value=active_formula(row))
    fc.font = F_FORMULA
    fc.fill = P_DERIVED
    fc.number_format = fmt
    fc.alignment = ALIGN_R
    ws.cell(row=row, column=7, value=tier).font = F_TIER
    ws.cell(row=row, column=8, value=status).font = F_TIER
    nc = ws.cell(row=row, column=9, value=notes); nc.font = F_NOTE
    nc.alignment = ALIGN_L


def write_text_input(ws, row, id_, desc, bear, base, bull, tier, status, notes,
                      flagged=False):
    """For text-valued assumptions (e.g., Terminal Method = 'Exit Multiple')."""
    ws.cell(row=row, column=1, value=id_).font = F_LABEL
    desc_cell = ws.cell(row=row, column=2, value=desc); desc_cell.font = F_LABEL
    desc_cell.alignment = ALIGN_L
    for col, val in zip([3, 4, 5], [bear, base, bull]):
        c = ws.cell(row=row, column=col, value=val)
        c.font = F_INPUT
        c.alignment = ALIGN_C
    fc = ws.cell(row=row, column=6, value=active_formula(row))
    fc.font = F_FORMULA
    fc.alignment = ALIGN_C
    ws.cell(row=row, column=7, value=tier).font = F_TIER
    sc = ws.cell(row=row, column=8, value=status)
    sc.font = F_FLAG if flagged else F_TIER
    nc = ws.cell(row=row, column=9, value=notes); nc.font = F_NOTE
    nc.alignment = ALIGN_L
    if flagged:
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = P_FLAG


# ── Build ──────────────────────────────────────────────────────────────────
def main():
    wb = load_workbook(WB_PATH)

    # Replace Assumptions sheet at same position
    old_idx = wb.sheetnames.index('Assumptions')
    del wb['Assumptions']
    ws = wb.create_sheet('Assumptions', old_idx)

    # Column widths
    widths = {'A': 7, 'B': 46, 'C': 12, 'D': 12, 'E': 12, 'F': 13,
              'G': 11, 'H': 14, 'I': 62}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:I1')
    t = ws['A1']
    t.value = 'WHOOP — Master Assumptions'
    t.font = F_TITLE
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Scenario toggle
    a3 = ws['A3']; a3.value = 'Scenario:'
    a3.font = F_HEADER; a3.fill = P_TOGGLE; a3.alignment = ALIGN_C
    a3.font = Font(name='Calibri', size=11, color='000000', bold=True)
    b3 = ws['B3']; b3.value = 'Base'
    b3.font = F_TOGGLE; b3.fill = P_TOGGLE; b3.alignment = ALIGN_C
    ws.merge_cells('C3:I3')
    c3 = ws['C3']
    c3.value = ('  Select Bear / Base / Bull — drives Active column F throughout '
                'the model. Other tabs reference Active values only.')
    c3.font = F_NOTE; c3.fill = P_TOGGLE
    c3.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 22

    dv = DataValidation(type='list', formula1='"Bear,Base,Bull"', allow_blank=False)
    dv.add('B3')
    ws.add_data_validation(dv)

    # Header row
    for i, h in enumerate(['ID', 'Assumption', 'Bear', 'Base', 'Bull',
                            'Active', 'Tier', 'Status', 'Notes'], 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = F_HEADER
        c.fill = P_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER_B
    ws.row_dimensions[5].height = 20

    # Freeze panes
    ws.freeze_panes = 'B6'

    # Row tracker for cross-references
    rmap = {}
    row = 7

    # ─── Section 1: Revenue & Growth ───────────────────────────────────────
    section_header(ws, row, 'Section 1: Revenue & Growth')
    row += 1

    rmap['R3a'] = row
    write_input(ws, row, 'R3a', '2025 Beginning Members (M)',
                1.5, 1.3, 1.1, 'Tier 3', 'High',
                'Back-solved to reconcile avg×ARPU ≈ $650M base. Bear slower decel; bull faster acquisition.',
                fmt='0.00')
    row += 1

    rmap['R3b'] = row
    write_input(ws, row, 'R3b', '2025 Ending Members (M)',
                2.3, 2.5, 2.7, 'Tier 1', 'Critical',
                'Series G press release: 2.5M+ paying members',
                fmt='0.00')
    row += 1

    rmap['R3'] = row
    write_derived(ws, row, 'R3', '2025 Average Members (M) — DERIVED',
                  {'C': f"=(C{rmap['R3a']}+C{rmap['R3b']})/2",
                   'D': f"=(D{rmap['R3a']}+D{rmap['R3b']})/2",
                   'E': f"=(E{rmap['R3a']}+E{rmap['R3b']})/2"},
                  'Derived', 'Critical',
                  'Per methodology: ARPU applies to AVERAGE members during period, NOT year-end.',
                  fmt='0.00')
    row += 1

    rmap['R4'] = row
    write_input(ws, row, 'R4', '2025 Blended ARPU ($/yr)',
                300, 338, 375, 'Tier 3', 'Critical',
                'v2 build per CLAUDE.md rebuild directives ($300/$338/$375). '
                'Single-line; 5-component decomposition deferred to ARPU tab.',
                fmt='$#,##0')
    row += 1

    rmap['R1'] = row
    write_derived(ws, row, 'R1', '2025 Recognized Revenue ($M) — DERIVED',
                  {'C': f"=C{rmap['R3']}*C{rmap['R4']}",
                   'D': f"=D{rmap['R3']}*D{rmap['R4']}",
                   'E': f"=E{rmap['R3']}*E{rmap['R4']}"},
                  'Derived', 'Critical',
                  'DERIVED: Avg Members × Blended ARPU. Target ~$650M base. '
                  'Methodology consistency check: must reconcile within ±5%.',
                  fmt='$#,##0')
    row += 1

    rmap['R2'] = row
    write_input(ws, row, 'R2', '2025 Bookings Run-Rate ($M)',
                1100, 1100, 1100, 'Tier 1', 'Locked',
                'Series G press release; CEO statement (Dec 2025 exit rate × 12)',
                fmt='$#,##0')
    row += 1

    rmap['R5'] = row
    write_input(ws, row, 'R5', 'Subscription Revenue %',
                0.80, 0.85, 0.88, 'Tier 2', 'High',
                'Sacra/BMCT: ~80-88% range', fmt='0.0%')
    row += 1

    rmap['R6'] = row
    write_input(ws, row, 'R6', '2025 YoY Subscription Growth',
                1.03, 1.03, 1.03, 'Tier 1', 'Locked',
                'Series G press release: 103%', fmt='0.0%')
    row += 1

    rmap['R7'] = row
    write_input(ws, row, 'R7', 'Phase 1 Revenue Growth (2026-27)',
                0.50, 0.70, 0.90, 'Tier 3', 'Critical',
                'Peer trajectories; Series G deployment & WHOOP 5.0 cycle', fmt='0.0%')
    row += 1

    rmap['R8'] = row
    write_input(ws, row, 'R8', 'Phase 2 Revenue Growth (2028-30)',
                0.15, 0.25, 0.40, 'Tier 3', 'Critical',
                'Deceleration shape: Dexcom (bull) / Netflix (base) / Peloton (bear)', fmt='0.0%')
    row += 1

    rmap['R9'] = row
    write_input(ws, row, 'R9', 'Phase 3 Revenue Growth (2031-33)',
                0.05, 0.12, 0.20, 'Tier 3', 'High',
                'TAM ceiling; healthcare attach resolution', fmt='0.0%')
    row += 1

    # Cohort churn replaces flat R10
    rmap['R10a'] = row
    write_input(ws, row, 'R10a', 'Year-1 Cohort Annual Churn',
                0.30, 0.22, 0.15, 'Tier 3', 'Critical',
                'Highest churn ("try and quit"). Will drive Members tab cohort engine.',
                fmt='0.0%', flagged=True)
    row += 1

    rmap['R10b'] = row
    write_input(ws, row, 'R10b', 'Year-2 Cohort Annual Churn',
                0.18, 0.14, 0.10, 'Tier 3', 'Critical',
                'Stabilizing cohort', fmt='0.0%', flagged=True)
    row += 1

    rmap['R10c'] = row
    write_input(ws, row, 'R10c', 'Year-3+ Cohort Annual Churn',
                0.12, 0.08, 0.05, 'Tier 3', 'Critical',
                'Retained core; very sticky power users', fmt='0.0%', flagged=True)
    row += 1

    rmap['R11'] = row
    write_input(ws, row, 'R11', 'International Revenue %',
                0.35, 0.43, 0.55, 'Tier 2', 'High',
                'Stress-tested: 40-45% base; NYT DealBook source ambiguous on "new sales" vs total',
                fmt='0.0%')
    row += 1

    rmap['R12'] = row
    write_input(ws, row, 'R12', 'B2B Unite Revenue %',
                0.12, 0.18, 0.25, 'Tier 2', 'Medium',
                'Sacra/BMCT: ~20% recurring revenue; growing rapidly',
                fmt='0.0%')
    row += 1

    rmap['R14'] = row
    write_input(ws, row, 'R14', 'Advanced Labs Attach Rate',
                0.03, 0.10, 0.25, 'Tier 3', 'Critical',
                'BUCKET-3 THESIS VARIABLE. Bear/Base/Bull per methodology Sec 3. '
                'Single most important ARPU assumption — swings $80/member.',
                fmt='0.0%', flagged=True)
    row += 1

    rmap['R15'] = row
    write_input(ws, row, 'R15', 'Labs Avg Test Price ($)',
                250, 300, 350, 'Tier 3', 'High',
                'WHOOP Advanced Labs site: $199-$599. Methodology base: $300.',
                fmt='$#,##0', flagged=True)
    row += 1

    rmap['R16'] = row
    write_input(ws, row, 'R16', 'Annual Price Inflation',
                0.02, 0.03, 0.04, 'Tier 3', 'Medium',
                'CPI + premium brand pricing power', fmt='0.0%')
    row += 1

    rmap['R18'] = row
    write_input(ws, row, 'R18', '2026 Gross Adds (M)',
                1.5, 1.9, 2.3, 'Tier 3', 'Critical',
                'Methodology priors: 1.8-2.0M base. Members tab will use this; per-year '
                'gross adds modeled there. CAC × this = S&M consistency check.',
                fmt='0.00', flagged=True)
    row += 1

    rmap['R19'] = row
    write_input(ws, row, 'R19', 'Deferred Revenue Adjustment (% of bookings)',
                -0.10, -0.08, -0.05, 'Tier 3', 'Medium',
                'Drag during growth (recognized < booked); ASC 606. Approaches 0 at steady state.',
                fmt='0.0%')
    row += 1

    row += 1  # spacer

    # ─── Section 2: Cost Structure & Margins ───────────────────────────────
    section_header(ws, row, 'Section 2: Cost Structure & Margins')
    row += 1

    rmap['C1'] = row
    write_input(ws, row, 'C1', 'Hardware BOM per Device ($)',
                100, 80, 60, 'Tier 3', 'Critical',
                'Est. $50-80 at scale; 2016 BOM was $250-300 per Contrary Research',
                fmt='$#,##0', flagged=True)
    row += 1

    rmap['C2'] = row
    write_input(ws, row, 'C2', 'Loaded Hardware Cost per Acquisition ($)',
                150, 120, 90, 'Tier 3', 'Critical',
                'C1 + assembly, ship, warranty, returns. Drives unit economics.',
                fmt='$#,##0', flagged=True)
    row += 1

    rmap['C4'] = row
    write_input(ws, row, 'C4', 'Subscription Delivery Cost (% of sub rev)',
                0.15, 0.10, 0.08, 'Tier 3', 'High',
                'Infra + support; AWS/GCP benchmarks; declining at scale', fmt='0.0%')
    row += 1

    rmap['C5'] = row
    write_input(ws, row, 'C5', 'Subscription Gross Margin (standalone)',
                0.72, 0.78, 0.83, 'Tier 3', 'Critical',
                'Peer: pure SaaS ~80%, Dexcom ~64% (incl consumables). WHOOP higher quality recurring.',
                fmt='0.0%')
    row += 1

    rmap['C6'] = row
    write_input(ws, row, 'C6', 'Hardware Gross Margin (standalone)',
                -0.10, -0.02, 0.05, 'Tier 3', 'Critical',
                'Subsidized at acquisition; Peloton HW GM trajectory: -18% to +14%',
                fmt='0.0%')
    row += 1

    rmap['C7a'] = row
    write_input(ws, row, 'C7a', 'Blended GM Phase 1 (2026-27)',
                0.56, 0.60, 0.64, 'Tier 3', 'Critical',
                'Derived prior; lower in growth years (more hardware drag)',
                fmt='0.0%')
    row += 1

    rmap['C7b'] = row
    write_input(ws, row, 'C7b', 'Blended GM Phase 2 (2028-30)',
                0.63, 0.67, 0.72, 'Tier 3', 'High',
                'Margin expansion as growth decelerates', fmt='0.0%')
    row += 1

    rmap['C7c'] = row
    write_input(ws, row, 'C7c', 'Blended GM Terminal',
                0.72, 0.76, 0.80, 'Tier 3', 'High',
                'Peer terminal: Dexcom 60%, Spotify 32%; software-heavy mix at scale',
                fmt='0.0%')
    row += 1

    rmap['C8a'] = row
    write_input(ws, row, 'C8a', 'R&D % of Revenue (current)',
                0.25, 0.22, 0.20, 'Tier 3', 'High',
                'Peers: Dexcom 13%, Garmin 16%, growth SaaS 20-30%', fmt='0.0%')
    row += 1

    rmap['C8b'] = row
    write_input(ws, row, 'C8b', 'R&D % of Revenue (terminal)',
                0.15, 0.13, 0.12, 'Tier 3', 'Medium',
                'Peer steady-state', fmt='0.0%')
    row += 1

    rmap['C9'] = row
    write_input(ws, row, 'C9', 'S&M % of Revenue (current)',
                0.35, 0.30, 0.25, 'Tier 3', 'Critical',
                'Peloton peak 28.4% FY22, current 16.9% FY25. Bear = stuck high.',
                fmt='0.0%')
    row += 1

    rmap['C9a'] = row
    write_input(ws, row, 'C9a', 'Paid CAC per Gross Add ($)',
                220, 175, 130, 'Tier 3', 'Critical',
                '30-50% of first-year ARPU. S&M consistency check on Checks tab.',
                fmt='$#,##0', flagged=True)
    row += 1

    rmap['C9b'] = row
    write_input(ws, row, 'C9b', 'Athlete/Brand Spend ($M/yr)',
                100, 60, 30, 'Tier 3', 'High',
                'Press, endorsement deals; non-paid S&M floor', fmt='$#,##0',
                flagged=True)
    row += 1

    rmap['C9t'] = row
    write_input(ws, row, 'C9t', 'S&M % of Revenue (terminal)',
                0.20, 0.17, 0.15, 'Tier 3', 'High',
                'Peer steady-state: Garmin SG&A 14-19%', fmt='0.0%')
    row += 1

    rmap['C10'] = row
    write_input(ws, row, 'C10', 'G&A % of Revenue (current)',
                0.18, 0.15, 0.12, 'Tier 3', 'Medium',
                'IPO step-up 200-300bps in 2027-28 (SOX, IR, legal)',
                fmt='0.0%')
    row += 1

    rmap['C10t'] = row
    write_input(ws, row, 'C10t', 'G&A % of Revenue (terminal)',
                0.12, 0.10, 0.08, 'Tier 3', 'Medium',
                'Peer steady-state', fmt='0.0%')
    row += 1

    rmap['C11'] = row
    write_input(ws, row, 'C11', 'SBC % of Revenue',
                0.15, 0.12, 0.10, 'Tier 3', 'High',
                'Peloton: 9.2% FY25; pre-IPO typically higher. SBC is REAL COST in FCF.',
                fmt='0.0%')
    row += 1

    rmap['C13'] = row
    write_input(ws, row, 'C13', 'CapEx % of Revenue',
                0.05, 0.04, 0.03, 'Tier 3', 'Medium',
                'Asset-light; contract manufacturing', fmt='0.0%')
    row += 1

    rmap['C16'] = row
    write_input(ws, row, 'C16', 'NWC % of Incremental Revenue',
                -0.05, -0.08, -0.10, 'Tier 3', 'High',
                'Negative = WC benefit (deferred rev > inventory drag during growth)',
                fmt='0.0%')
    row += 1

    rmap['C17'] = row
    write_input(ws, row, 'C17', 'Advanced Labs COGS %',
                0.70, 0.65, 0.60, 'Tier 3', 'High',
                'Quest/Labcorp panel margins ~30-40%; bucket-3 sensitivity',
                fmt='0.0%')
    row += 1

    rmap['C15'] = row
    write_input(ws, row, 'C15', 'Tax Rate (effective)',
                0.25, 0.23, 0.21, 'Tier 2', 'Medium',
                'US statutory; intl mix-shift lowers over time', fmt='0.0%')
    row += 1

    row += 1

    # ─── Section 3: Capital Structure & Cap Table ──────────────────────────
    section_header(ws, row, 'Section 3: Capital Structure & Cap Table')
    row += 1

    rmap['S1'] = row
    write_input(ws, row, 'S1', 'Fully Diluted Shares Outstanding (M)',
                420, 410, 400, 'Tier 3', 'Critical',
                '361.2M common + ~50M option pool est. Pitchbook auth pref 537.9M, outs 361.2M.',
                fmt='#,##0', flagged=True)
    row += 1

    rmap['S2'] = row
    write_input(ws, row, 'S2', 'Series G Price per Share ($)',
                28.00, 28.00, 28.00, 'Tier 2', 'High',
                'Pitchbook Series G data; back-solve from $575M raise / shares issued',
                fmt='$#,##0.00', flagged=True)
    row += 1

    rmap['S3'] = row
    write_text_input(ws, row, 'S3', 'Liquidation Preference Type',
                'Participating', '1x non-participating', '1x non-participating',
                'Tier 2', 'Locked',
                'Pitchbook: 1x non-participating pari passu. Bear assumes participating ratchet.')
    row += 1

    rmap['S4'] = row
    write_input(ws, row, 'S4', 'Total Liquidation Preference ($M)',
                1400, 1400, 1400, 'Tier 2', 'High',
                'Sum of preferred raises: ~$1.4B aggregate (Series A through G)',
                fmt='$#,##0')
    row += 1

    rmap['S7'] = row
    write_input(ws, row, 'S7', 'Option Pool Size (% of FD shares)',
                0.15, 0.12, 0.10, 'Tier 3', 'High',
                'Typical 10-15% pre-IPO; 600-person hiring surge implies aggressive granting',
                fmt='0.0%', flagged=True)
    row += 1

    rmap['S8'] = row
    write_input(ws, row, 'S8', 'Founder Ownership Ahmed (% FD)',
                0.10, 0.14, 0.18, 'Tier 3', 'Medium',
                '12-18% est. across 7 rounds + pool refreshes; possibly lower with secondaries',
                fmt='0.0%', flagged=True)
    row += 1

    rmap['S9'] = row
    write_input(ws, row, 'S9', 'DLOM (Discount for Lack of Marketability)',
                0.30, 0.22, 0.15, 'Tier 2', 'High',
                'Oura tender 25%; EquityZen avg 13-29%; Hiive 39-51% deeper discount',
                fmt='0.0%')
    row += 1

    rmap['S10'] = row
    write_input(ws, row, 'S10', 'Debt on Balance Sheet ($M)',
                50, 10, 0, 'Tier 2', 'Low',
                'Likely minimal/zero; Pitchbook shows $10.13M debt round Mar 2024',
                fmt='$#,##0')
    row += 1

    rmap['S11'] = row
    write_input(ws, row, 'S11', 'Cash post Series G ($M)',
                500, 700, 850, 'Tier 3', 'Medium',
                'Back-solve: $575M raise + operating cash flow positive entering 2026',
                fmt='$#,##0', flagged=True)
    row += 1

    rmap['S12'] = row
    write_derived(ws, row, 'S12', 'EV-to-Equity Bridge ($M) — DERIVED',
                  {'C': f"=C{rmap['S11']}-C{rmap['S10']}",
                   'D': f"=D{rmap['S11']}-D{rmap['S10']}",
                   'E': f"=E{rmap['S11']}-E{rmap['S10']}"},
                  'Derived', 'High',
                  'Cash − Debt: add to EV to bridge to equity value',
                  fmt='$#,##0')
    row += 1

    rmap['S15'] = row
    write_text_input(ws, row, 'S15', 'Series F Ratchet / IPO Provisions',
                'Live', 'Likely extinguished', 'Extinguished',
                'Tier 3', 'High',
                'Likely extinguished at Series G close; SoftBank notably absent from Series G syndicate',
                flagged=True)
    row += 1

    rmap['S16'] = row
    write_text_input(ws, row, 'S16', 'Dual-Class Share Structure (IPO)',
                'Single class', 'Dual class', 'Dual class',
                'Tier 3', 'Medium',
                'Common for founder-led tech IPOs (Meta, Snap, Airbnb); not disclosed',
                flagged=True)
    row += 1

    row += 1

    # ─── Section 4: WACC / Discount Rate ───────────────────────────────────
    section_header(ws, row, 'Section 4: WACC / Discount Rate (Damodaran Apr 2026)')
    row += 1

    rmap['D1'] = row
    write_input(ws, row, 'D1', 'Risk-Free Rate (10yr UST)',
                0.043, 0.043, 0.043, 'Tier 1', 'Locked',
                'Fed H.15, Apr 14 2026: 4.26%; rounded 4.3%', fmt='0.00%')
    row += 1

    rmap['D2'] = row
    write_input(ws, row, 'D2', 'Equity Risk Premium',
                0.045, 0.045, 0.045, 'Tier 1', 'Point Est.',
                'Damodaran 4.23% (Jan 2026); Kroll 5.0%; midpoint 4.5%', fmt='0.00%')
    row += 1

    rmap['D3'] = row
    write_input(ws, row, 'D3', 'Unlevered Beta (bucket-weighted)',
                0.85, 0.83, 0.81, 'Tier 1', 'Sensitivity',
                'Damodaran sector: Healthcare 0.83, Electronics 0.83. Bucket-weighted.',
                fmt='0.00')
    row += 1

    rmap['D4'] = row
    write_input(ws, row, 'D4', 'Size Premium',
                0.0081, 0.0081, 0.0081, 'Tier 2', 'Point Est.',
                'Kroll CRSP Decile 3 ($7.3-13.5B): 0.81%', fmt='0.00%')
    row += 1

    rmap['D5'] = row
    write_input(ws, row, 'D5', 'Company-Specific Risk Premium (CSRP)',
                0.030, 0.020, 0.010, 'Tier 3', 'Sensitivity',
                'Pre-IPO execution risk + private illiquidity; 1-3% range', fmt='0.00%')
    row += 1

    rmap['D8'] = row
    write_derived(ws, row, 'D8', 'Implied Cost of Equity / WACC — DERIVED',
                  {'C': f"=C{rmap['D1']}+C{rmap['D2']}*C{rmap['D3']}+C{rmap['D4']}+C{rmap['D5']}",
                   'D': f"=D{rmap['D1']}+D{rmap['D2']}*D{rmap['D3']}+D{rmap['D4']}+D{rmap['D5']}",
                   'E': f"=E{rmap['D1']}+E{rmap['D2']}*E{rmap['D3']}+E{rmap['D4']}+E{rmap['D5']}"},
                  'Derived', 'Critical',
                  '= D1 + D2×D3 + D4 + D5; CAPM build, no debt component (100% equity)',
                  fmt='0.00%')
    row += 1

    row += 1

    # ─── Section 5: Terminal Value ─────────────────────────────────────────
    section_header(ws, row, 'Section 5: Terminal Value')
    row += 1

    rmap['D9'] = row
    write_text_input(ws, row, 'D9', 'Terminal Method (primary)',
                'Exit Multiple', 'Exit Multiple', 'Exit Multiple',
                'Design', 'Locked',
                'Primary: EV/Revenue exit multiple; cross-check: Gordon Growth')
    row += 1

    rmap['D11'] = row
    write_input(ws, row, 'D11', 'Terminal EV/Revenue Multiple',
                3.5, 4.5, 6.0, 'Tier 3', 'Critical',
                'Compressed 40-60% from current comps; reflects terminal growth 10-15%',
                fmt='0.0"x"')
    row += 1

    rmap['D12'] = row
    write_input(ws, row, 'D12', 'Terminal EV/EBITDA Multiple',
                12, 15, 18, 'Tier 3', 'High',
                'Cross-check only; more stable than revenue multiple at maturity',
                fmt='0.0"x"')
    row += 1

    rmap['D13'] = row
    write_input(ws, row, 'D13', 'Perpetuity Growth Rate (Gordon)',
                0.025, 0.030, 0.035, 'Tier 2', 'Critical',
                'GDP growth + inflation; cannot exceed WACC', fmt='0.0%')
    row += 1

    rmap['D28'] = row
    write_input(ws, row, 'D28', 'Multiple Compression Factor',
                0.40, 0.50, 0.60, 'Tier 3', 'High',
                'Terminal multiples 40-60% lower than current comps. Applied to derive D11.',
                fmt='0.0%')
    row += 1

    row += 1

    # ─── Section 6: Comp Bucket Weighting (THE THESIS) ─────────────────────
    section_header(ws, row, 'Section 6: Comp Bucket Weighting — THIS IS THE THESIS')
    row += 1

    rmap['V11a'] = row
    write_input(ws, row, 'V11a', 'Bucket 1 Weight (Consumer Hardware — Garmin)',
                0.40, 0.20, 0.10, 'Design', 'Critical',
                'Higher = more bearish (WHOOP as hardware company)', fmt='0.0%')
    row += 1

    rmap['V11b'] = row
    write_input(ws, row, 'V11b', 'Bucket 2 Weight (Consumer Subscription — PTON/SPOT)',
                0.40, 0.40, 0.30, 'Design', 'Critical',
                'Peloton-Spotify range disaggregated; where does WHOOP sit?',
                fmt='0.0%')
    row += 1

    rmap['V11c'] = row
    write_input(ws, row, 'V11c', 'Bucket 3 Weight (Health Data — DXCM/RMD/MASI/IRTC)',
                0.20, 0.40, 0.60, 'Design', 'Critical',
                'Higher = healthcare thesis is believed; Series G implied ~70%',
                fmt='0.0%')
    row += 1

    rmap['V11_chk'] = row
    write_derived(ws, row, 'V11_chk', 'Bucket Weights Sum (must = 100%)',
                  {'C': f"=C{rmap['V11a']}+C{rmap['V11b']}+C{rmap['V11c']}",
                   'D': f"=D{rmap['V11a']}+D{rmap['V11b']}+D{rmap['V11c']}",
                   'E': f"=E{rmap['V11a']}+E{rmap['V11b']}+E{rmap['V11c']}"},
                  'Derived', 'Check',
                  'Sanity check: bucket weights must sum to 100%', fmt='0.0%')
    row += 1

    row += 1

    # ─── Section 7: IPO / Discount Framework ────────────────────────────────
    section_header(ws, row, 'Section 7: IPO Trading Range & Discount Framework')
    row += 1

    rmap['I1'] = row
    write_input(ws, row, 'I1', 'IPO Timing (months from Series G)',
                18, 15, 12, 'Tier 2', 'High',
                'Q2 2027 - Q1 2028; "last private round" per Ahmed', fmt='0')
    row += 1

    rmap['I2'] = row
    write_input(ws, row, 'I2', 'IPO Pricing Discount (to Stage 1 fair value)',
                0.25, 0.20, 0.15, 'Tier 2', 'Critical',
                'Loughran/Ritter IPO pricing studies; 15-25% standard', fmt='0.0%')
    row += 1

    rmap['I3'] = row
    write_input(ws, row, 'I3', 'Illiquidity / Timing Discount (private vs. IPO)',
                0.15, 0.10, 0.05, 'Tier 2', 'High',
                'Oura tender 25%; Hiive 39-51%; pre-IPO DLOM shrinkage near window',
                fmt='0.0%')
    row += 1

    rmap['I4'] = row
    write_input(ws, row, 'I4', 'Series G Investor Required IRR',
                0.30, 0.25, 0.20, 'Tier 3', 'Medium',
                'Late-stage growth equity return expectations (3-5yr hold)',
                fmt='0.0%')
    row += 1

    rmap['I5'] = row
    write_derived(ws, row, 'I5', 'Implied Stage 1 Public Fair Value ($M) — DERIVED',
                  {'C': f"=10100/(1-C{rmap['I3']})/(1-C{rmap['I2']})",
                   'D': f"=10100/(1-D{rmap['I3']})/(1-D{rmap['I2']})",
                   'E': f"=10100/(1-E{rmap['I3']})/(1-E{rmap['I2']})"},
                  'Derived', 'Critical',
                  'Back-solve: $10.1B / (1-illiquidity) / (1-IPO discount) → $13-15B target',
                  fmt='$#,##0')
    row += 1

    row += 1

    # ─── Section 8: Precedent Transactions ──────────────────────────────────
    section_header(ws, row, 'Section 8: Precedent Transactions')
    row += 1

    rmap['P1'] = row
    write_text_input(ws, row, 'P1', 'Deal Inclusion Window',
                'Last 5 years', 'Last 5 years', 'Last 5 years',
                'Design', 'Locked',
                '2021-2026 with 3-year primary focus')
    row += 1

    rmap['P3'] = row
    write_input(ws, row, 'P3', 'Deal Size Threshold ($M EV)',
                500, 500, 500, 'Design', 'Locked',
                'Below this, synergy math distorts multiples', fmt='$#,##0')
    row += 1

    rmap['P5'] = row
    write_text_input(ws, row, 'P5', 'Multiple Extraction Method',
                'EV/LTM Rev', 'EV/LTM Rev', 'EV/LTM Rev',
                'Design', 'Locked',
                'EV / LTM revenue at announcement; NTM where available')
    row += 1

    rmap['P_b1'] = row
    write_input(ws, row, 'P_b1', 'Bucket 1 Precedent Multiple (median)',
                1.0, 1.5, 2.5, 'Tier 2', 'High',
                'Fitbit/Google ~1.3x; Garmin tuck-ins; mostly distressed exits',
                fmt='0.0"x"')
    row += 1

    rmap['P_b2'] = row
    write_input(ws, row, 'P_b2', 'Bucket 2 Precedent Multiple (median)',
                0.8, 1.5, 3.0, 'Tier 2', 'High',
                'Peloton/Precor; connected fitness all value-destructive',
                fmt='0.0"x"')
    row += 1

    rmap['P_b3'] = row
    write_input(ws, row, 'P_b3', 'Bucket 3 Precedent Multiple (median)',
                6.0, 8.0, 10.0, 'Tier 2', 'High',
                'Per research/precedent-transactions.md: B3 medtech 6.6-10.1x range',
                fmt='0.0"x"')
    row += 1

    row += 1

    # ─── Section 9: Real Options ───────────────────────────────────────────
    section_header(ws, row, 'Section 9: Real Options — moved to RealOptions tab (additive, never in DCF base case)')
    row += 1
    note_cell = ws.cell(row=row, column=1,
                        value=('Real-option probabilities (D19-D23) live on the RealOptions tab — '
                               'cluster-modeled (correlated success / mixed / failure), additive '
                               'to base DCF EV. See methodology Sec 4.'))
    note_cell.font = F_NOTE
    note_cell.alignment = ALIGN_L
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2

    # ─── Legend ─────────────────────────────────────────────────────────────
    section_header(ws, row, 'Legend & Conventions')
    row += 1
    legend = [
        ('Blue text',  'Hard-coded input',                                    F_INPUT,   None),
        ('Black text', 'Formula / derived calculation',                       F_FORMULA, P_DERIVED),
        ('Green text', 'Cross-sheet link (wired when other tabs are built)',  F_LINK,    None),
        ('Yellow fill','FLAGGED_GAP — unsourced critical, needs verification',F_FORMULA, P_FLAG),
        ('Active col', 'Picks Bear/Base/Bull per scenario selector at $B$3',  F_FORMULA, None),
    ]
    for label, desc, font, fill in legend:
        a = ws.cell(row=row, column=1, value=label); a.font = font
        if fill: a.fill = fill
        a.alignment = ALIGN_C
        b = ws.cell(row=row, column=2, value=desc); b.font = F_NOTE
        b.alignment = ALIGN_L
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1

    # Save
    wb.save(WB_PATH)
    print(f'Wrote {WB_PATH}')
    print(f'Final row used: {row - 1}')
    print(f'Assumption count: {len(rmap)}')


if __name__ == '__main__':
    main()
